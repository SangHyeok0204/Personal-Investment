"""[금리 5주제] — AI Key Data `금리_2.xlsx` 판독 (2026-08-28).

신상품팀 202608 공모손차 데이터를 AI Key Data 프로젝트로 복사해 온 워크북이다.
정본은 이 파일 하나이고(같은 폴더의 나머지 10개는 이 워크북을 만든 블룸버그 원천),
주제 5개가 `가공시트 + _raw_data` 쌍으로 들어 있다. 계약서는 같은 폴더의 `_출처.md`.

  금리(1) 미 하이퍼스케일러 채권 발행   221건 · 2020~2026-08
  금리(2) 미국 인플레이션 지표         CPI YoY · 클리블랜드 기대 · 1Y 스왑 · 트루플레이션
  금리(3) WTI 유가                    CL1 · CL6 · CL12 + CL1−CL12 스프레드
  금리(4) ADP 민간고용                 월별 증감 + 12개월 이동평균
  금리(5) FOMC 금리인상 내재확률       일별 %

왜 한 모듈·한 엔드포인트인가 — 다섯 카드가 **같은 파일 한 장**을 본다. 주제마다
모듈을 두면 같은 1.6MB 워크북을 다섯 번 연다(SMB 왕복이 그만큼 는다).

★일별 시계열(금리2·3)은 3,886행이라 그대로 내보내면 payload 가 1MB 를 넘는다.
  카드 폭이 몇백 px 인데 그 해상도가 화면에 남지도 않으므로 **주간 마지막값**으로
  솎아 낸다(macro_panels 의 _weekly 와 같은 접근).

⚠️채권 발행액은 **발행 통화 액면을 그대로 합산**한 값이다(워크북의 Year 요약 열이
  그렇게 만들어져 있고 리포트 서술도 그 숫자를 쓴다). USD 가 $400B 로 대부분이지만
  EUR·CAD·GBP·CHF·AUD·JPY 가 섞여 있어 엄밀한 달러 환산이 아니다.
"""
from __future__ import annotations

import io
import os
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone

_KST = timezone(timedelta(hours=9))

SRC_PATH = os.environ.get(
    "RATE_TOPICS_XLSX", "/srv/legacy/gpu_compute/금리/금리_2.xlsx"
)

# 시트는 이름 전체가 아니라 `금리(N)` 접두로 찾는다 — 원본 시트명에 오타가 있고
# ("인프레이션") 누가 고칠 수 있다. `_raw`/`_raw_data` 짝은 제외한다(가공 시트만 쓴다).
TOPIC_PREFIX = {1: "금리(1)", 2: "금리(2)", 3: "금리(3)", 4: "금리(4)", 5: "금리(5)"}

HEADER_ROW = 4  # 1-based: r4 가 컬럼명, r5 부터 데이터 (B열부터 시작)


def _num(v) -> float | None:
    """숫자만 통과. 시트에 '#N/A' 문자열이 섞여 있다."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _as_date(v) -> date | None:
    """★채권 시트의 `Issue Date`·`Maturity` 는 **텍스트**로 저장돼 있다
    (`_출처.md` 가 경고한 항목). 날짜 셀과 'YYYY-MM-DD' 문자열을 둘 다 받는다."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return datetime.strptime(v.strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _weekly(points: list[tuple[date, float]]) -> list[list]:
    """주(월요일 기준) 마지막 관측치만 남긴다. 날짜 오름차순 입력 전제."""
    keep: dict[tuple[int, int], tuple[date, float]] = {}
    for d, v in points:
        iso = d.isocalendar()
        keep[(iso[0], iso[1])] = (d, v)
    return [[d.isoformat(), v] for d, v in sorted(keep.values())]


# ── 주제별 파서 (시트 행 → payload 조각) ─────────────────────────────────────

def parse_bonds(rows: list[tuple]) -> dict:
    """금리(1) — B~K 가 발행 내역, M~N 이 워크북이 미리 접어 둔 연도별 합계."""
    by_year: list[list] = []
    by_issuer: dict[str, dict] = {}
    latest: date | None = None
    n = 0
    for r in rows:
        # 연도 요약(M=12, N=13) — 행마다 있는 게 아니라 위쪽 몇 행에만 있다.
        if len(r) > 13:
            y, amt = _num(r[12]), _num(r[13])
            if y and amt is not None:
                by_year.append([int(y), round(amt, 3)])
        name, ticker, issued, amt = (
            r[1] if len(r) > 1 else None,
            r[2] if len(r) > 2 else None,
            _as_date(r[3]) if len(r) > 3 else None,
            _num(r[4]) if len(r) > 4 else None,
        )
        if not name or amt is None:
            continue
        n += 1
        key = str(ticker or name)
        g = by_issuer.setdefault(key, {"ticker": key, "name": str(name), "amt_b": 0.0, "n": 0})
        g["amt_b"] += amt / 1e9
        g["n"] += 1
        if issued and (latest is None or issued > latest):
            latest = issued

    issuers = sorted(by_issuer.values(), key=lambda g: -g["amt_b"])
    for g in issuers:
        g["amt_b"] = round(g["amt_b"], 2)
    by_year.sort()
    return {
        "by_year": by_year,
        "by_issuer": issuers,
        "total_b": round(sum(a for _, a in by_year), 2),
        "n": n,
        "asof": latest.isoformat() if latest else None,
        "unit": "십억(발행 통화 액면 합산)",
    }


def _multi_series(rows: list[tuple], header: tuple, keys: list[str], weekly: bool) -> dict:
    """날짜 1열 + 값 N열 시트 → {asof, series:[{key,label,last,points}]}."""
    cols = [(i, str(header[i]).strip()) for i in range(2, len(header)) if header[i]]
    series: list[dict] = []
    asof: date | None = None
    for slot, (ci, label) in enumerate(cols):
        pts: list[tuple[date, float]] = []
        for r in rows:
            d = _as_date(r[1]) if len(r) > 1 else None
            v = _num(r[ci]) if len(r) > ci else None
            if d is None or v is None:
                continue
            pts.append((d, v))
        if not pts:
            continue
        pts.sort()
        if asof is None or pts[-1][0] > asof:
            asof = pts[-1][0]
        series.append({
            "key": keys[slot] if slot < len(keys) else f"s{slot}",
            "label": label,
            "last": pts[-1][1],
            "last_date": pts[-1][0].isoformat(),
            "points": _weekly(pts) if weekly else [[d.isoformat(), v] for d, v in pts],
        })
    return {"asof": asof.isoformat() if asof else None, "series": series}


def parse_inflation(rows, header) -> dict:
    return _multi_series(rows, header, ["cpi", "cleveland", "swap1y", "truflation"], True)


def parse_wti(rows, header) -> dict:
    return _multi_series(rows, header, ["cl1", "cl6", "cl12", "spread"], True)


def parse_adp(rows, header) -> dict:
    # 월별이라 솎지 않는다(127행).
    out = _multi_series(rows, header, ["chg", "ma12"], False)
    return out


def parse_fomc_prob(rows, header) -> dict:
    # 60행 일별.
    return _multi_series(rows, header, ["prob"], False)


def build_payload(sheets: dict[int, tuple[tuple, list[tuple]]]) -> dict:
    """{주제번호: (헤더행, 데이터행들)} → 카드 payload."""
    out: dict = {
        "generated_at": datetime.now(_KST).strftime("%Y-%m-%d %H:%M:%S"),
        "note": None,
        "bonds": None,
        "inflation": None,
        "wti": None,
        "adp": None,
        "fomc_prob": None,
    }
    if 1 in sheets:
        out["bonds"] = parse_bonds(sheets[1][1])
    if 2 in sheets:
        out["inflation"] = parse_inflation(sheets[2][1], sheets[2][0])
    if 3 in sheets:
        out["wti"] = parse_wti(sheets[3][1], sheets[3][0])
    if 4 in sheets:
        out["adp"] = parse_adp(sheets[4][1], sheets[4][0])
    if 5 in sheets:
        out["fomc_prob"] = parse_fomc_prob(sheets[5][1], sheets[5][0])
    return out


# ── xlsx 판독 (mtime+size 캐시) ──────────────────────────────────────────────
# 1.6MB · 6,700행짜리 워크북이라 매 요청 파싱은 수 초. compute_index 와 같은 idiom.
_CACHE: dict = {"sig": None, "sheets": None}


def _read_sheets(path: str = SRC_PATH) -> dict[int, tuple[tuple, list[tuple]]]:
    st = os.stat(path)
    sig = (st.st_mtime_ns, st.st_size)
    if _CACHE["sig"] == sig and _CACHE["sheets"] is not None:
        return _CACHE["sheets"]

    with open(path, "rb") as f:
        blob = f.read()
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
        sheets: dict[int, tuple[tuple, list[tuple]]] = {}
        for num, prefix in TOPIC_PREFIX.items():
            name = next(
                (s for s in wb.sheetnames
                 if s.startswith(prefix) and not s.endswith(("_raw", "_raw_data"))),
                None,
            )
            if not name:
                continue
            ws = wb[name]
            rows = list(ws.iter_rows(min_row=HEADER_ROW, values_only=True))
            if rows:
                sheets[num] = (rows[0], rows[1:])
        wb.close()
    except Exception:
        if _CACHE["sheets"] is not None:
            return _CACHE["sheets"]
        raise

    _CACHE["sig"] = sig
    _CACHE["sheets"] = sheets
    return sheets


def build_rate_topics() -> dict:
    """xlsx → 카드 payload 한 장. 원천 결측은 503 이 아니라 note 로 알린다."""
    try:
        sheets = _read_sheets()
    except FileNotFoundError:
        out = build_payload({})
        out["note"] = f"원천 파일이 없습니다 — {SRC_PATH}"
        return out
    return build_payload(sheets)
