"""[국내상장 ETF 자금·수익률] — `국내상장ETF 모니터링.xlsm` 의 `value` 시트 판독 (2026-09-01).

장 마감 뒤 "개인 자금이 어느 분류로 몰렸고 그 분류의 수익률은 어땠나" 를 보는 페이지의
데이터원. 원천은 운용역이 매일 굽는 워크북 한 장뿐이다(S: 의 매크로가 CT/CTD 로 채운다).

★★단위·의미는 시트 수식에서 확정했다(추정 아님):
    E열 거래대금(억)   = CT(code, 16002)/1e8          — 원 → 억
    F열 개인 순매수(억) = CT(code, "6511^10")/1e5      — 천원 → 억  (당일 개인 순매수)
    O~R MMT_1w..6m    = CTD("RATE", code, 시작, 끝, "15001")/100   — 기간 수익률(소수)
    S~V FF_1w..6m     = CTD("SUM",  code, 시작, 끝, "6511^10")/1e5 — **같은 6511^10 의
                         기간 합계**. 즉 FF 는 "펀드 자금유입"이 아니라 **개인 순매수 누적**이다.
  ⚠️이름만 보고 FF 를 설정/환매(자금유입)로 읽으면 페이지 전체가 틀린 말을 하게 된다.
  검증(2026-08-31 기준일): 당일 F열 상·하위가 daily_analysis/20260831.txt 의
  "개인순매수 상위/하위 10개 ETF" 와 종목·금액 모두 일치했다.

★★기간 창은 시트 1·2행(15~22열)에 **명시**돼 있다 — 우리가 날짜를 세지 않는다.
  20260831 기준: 1w=0824~0831, 1m=0731~0831, 3m=0531~0831, 6m=0228~0831.
  네 창은 전부 **끝일이 같고 오늘을 포함**한다(당일 ⊂ 1주 ⊂ 1달 ⊂ 3달 ⊂ 6달).
  그래서 "구간 분해"(1주~1달, 1~3달, 3~6달)는 누적끼리 빼서 만든다 — 겹침을 안 걷어내면
  같은 돈을 네 번 센다.

★★수익률 집계는 **ETF 단위로 먼저 구간 수익률을 만들고**(복리 체인) 그 다음 가중한다.
  누적 평균끼리 나누면 분류가 바뀔 때 종목 구성이 달라져 값이 미끄러진다.
  가중치는 **현재 시총**이다 — 과거 구간에 현재 시총을 쓰는 근사임을 화면에 적어 둔다.
  단순평균도 같이 낸다(daily_analysis txt 의 "중분류 평균" 과 맞춰 보기 위해서다).

★★절대 억원만 보면 순위가 늘 시장형/S&P500 이다(규모가 이긴다). 그래서 강도
  `net/시총`(%)을 같이 낸다 — 주간가격모니터가 "저점 대비 픽" 에서 배운 것과 같은 교훈.

★★워크북은 매일 **덮어쓰기**라 그 안에 과거가 없다. HISTORICAL 은 두 갈래로 만든다:
  1) 구간 분해 — 오늘 스냅샷 하나로 4구간을 만든다. **첫날부터 그려진다.**
  2) 시점별 추이 — 스냅샷을 sqlite 에 적재해 쌓는다(read-through + 배경 루프, 기준일 멱등).
     원천 폴더의 백업본까지 훑어(`seed_archive`) 2026-04·05·06·08 의 7시점을 복원했다.
     ★관측이 성기므로 "일별 누적"이 아니라 **각 시점에서 본 그 기간의 값**을 그린다 —
       관측 하나가 이미 누적이라 날짜가 띄엄띄엄해도 정확하다. 일별이 쌓이면 저절로 촘촘해진다.

★★워크북이 결측 대신 0 을 주는 자리가 있다(`_valid_return` 참조). CTD("RATE")는 창 시작에
  종목이 없으면 0.0 이다 — 그대로 쓰면 "3개월간 0% 였다"는 거짓 문장이 되고 신규 대형
  ETF 가 분류 평균을 0 으로 끌어내린다. 판독 직후 한 번 걸러서 화면·적재·이력이 같은 값을 본다.
"""
from __future__ import annotations

import io
import os
import re
import sqlite3
from datetime import date, datetime, timedelta, timezone

_KST = timezone(timedelta(hours=9))

SRC_PATH = os.environ.get(
    "ETF_CLASS_XLSM",
    "/srv/legacy/etf_monitor/국내상장ETF 모니터링.xlsm",
)
DB_PATH = os.environ.get("ETF_CLASS_DB", "/srv/storage/etf_monitor/history.sqlite")
SHEET = "value"
HEADER_ROW = 5      # 5행 = 컬럼명, 6행부터 데이터
ASOF_CELL = (4, 1)  # A4 = 기준일 'YYYYMMDD'

# 시트 열 인덱스(0-base)와 기대 헤더. 열이 밀리면 조용히 틀린 값을 낸다 — 그래서 대조한다.
COLS = {
    "code": (1, "종목코드"),
    "name": (2, "ETF명"),
    "listed": (3, "상장일"),
    "amt": (4, "거래대금(억)"),
    "net": (5, "개인 순매수(억)"),
    "chg": (6, "등락률"),
    "price": (7, "현재가(원)"),
    "mcap": (8, "시총(억)"),
    "country": (9, "투자 국가"),
    "gubun": (10, "구분"),
    "big": (11, "대분류"),
    "mid": (12, "중분류"),
    "small": (13, "소분류"),
    "mmt_1w": (14, "MMT_1w"),
    "mmt_1m": (15, "MMT_1m"),
    "mmt_3m": (16, "MMT_3m"),
    "mmt_6m": (17, "MMT_6m"),
    "ff_1w": (18, "FF_1w"),
    "ff_1m": (19, "FF_1m"),
    "ff_3m": (20, "FF_3m"),
    "ff_6m": (21, "FF_6m"),
}
# 나중에 생긴 열. 없어도 계산에 쓰이는 값이 아니라 **표시용 플래그**라 판독을 거부하지
# 않는다 — 이걸 필수로 두면 2026-04·05 백업본이 통째로 버려지고 이력이 5개월 짧아진다.
OPTIONAL_COLS = {
    "interest": (22, "관심ETF여부"),
}
# 그 열이 없는 워크북에서 쓸 기본값.
OPTIONAL_DEFAULT = {"interest": False}
_HORIZONS = ("1w", "1m", "3m", "6m")

# 화면이 고르는 기간. 누적(창 그대로) 다섯 + 구간(누적 차이) 넷.
PERIODS = [
    {"key": "d",  "label": "당일",  "span": None},
    {"key": "1w", "label": "1주",   "span": "1w"},
    {"key": "1m", "label": "1개월", "span": "1m"},
    {"key": "3m", "label": "3개월", "span": "3m"},
    {"key": "6m", "label": "6개월", "span": "6m"},
]
INTERVALS = [
    {"key": "1w", "label": "최근 1주",  "outer": "1w", "inner": None},
    {"key": "1m", "label": "1주~1개월", "outer": "1m", "inner": "1w"},
    {"key": "3m", "label": "1~3개월",   "outer": "3m", "inner": "1m"},
    {"key": "6m", "label": "3~6개월",   "outer": "6m", "inner": "3m"},
]

# 분류 축. path 는 트리 조상 열 — 표를 접었다 폈다 하는 데 쓴다.
AXES = [
    {"key": "gubun",   "label": "구분",     "col": "gubun",   "path": []},
    {"key": "big",     "label": "대분류",   "col": "big",     "path": ["gubun"]},
    {"key": "mid",     "label": "중분류",   "col": "mid",     "path": ["gubun", "big"]},
    {"key": "small",   "label": "소분류",   "col": "small",   "path": ["gubun", "big", "mid"]},
    {"key": "country", "label": "투자국가", "col": "country", "path": []},
]
_AXIS_BY_KEY = {a["key"]: a for a in AXES}
DEFAULT_AXIS = "mid"

_PERIOD_KEYS = [p["key"] for p in PERIODS]
_IV_KEYS = [s["key"] for s in INTERVALS]


# ── 값 다루기 ────────────────────────────────────────────────────────────────

def _num(v) -> float | None:
    """엑셀 셀 → float. '#VALUE!' 같은 오류 문자열·bool 은 결측으로 본다."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _txt(v) -> str:
    return str(v).strip() if v not in (None, "") else ""


def _asof_date(raw) -> str | None:
    """'YYYYMMDD'(또는 날짜형) → 'YYYY-MM-DD'."""
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    s = re.sub(r"\D", "", str(raw or ""))
    if len(s) != 8:
        return None
    try:
        return date(int(s[:4]), int(s[4:6]), int(s[6:])).isoformat()
    except ValueError:
        return None


def _fold(label: str) -> str:
    """분류 라벨의 **표기 차이**만 걷어낸 묶음 키. 뜻이 같은데 철자가 달라 갈리는 걸 막는다.

    실측(2026-09-01, 스냅샷 7시점): 중분류 `MSCI KOREA`/`MSCI Korea`, 소분류
    `Top10`/`TOP10`(한 워크북 안에서도 둘 다 나온다) · `AI SW 전반`/`AI SW전반` 등 14건.
    접지 않으면 시계열이 두 줄로 갈리고 표에서도 같은 분류가 두 행이 된다.
    ★표시는 접은 결과가 아니라 **원래 철자**로 한다 — 회의에서 쓰는 말이 바뀌면 안 된다.
    """
    return label.strip().casefold().replace(" ", "")


def _listed_before(listed: str, window_start: str | None) -> bool:
    """그 창이 시작될 때 이 ETF 가 이미 상장돼 있었나. 날짜가 없으면 판단하지 않는다(True)."""
    if not window_start or not listed:
        return True
    return str(listed)[:10] <= window_start


def _valid_return(v: float | None, listed: str, window_start: str | None) -> float | None:
    """기간 수익률을 그대로 쓸지, '관측 없음'으로 볼지.

    ★★워크북의 CTD("RATE") 는 창 시작 시점에 종목이 없으면 **0.0 을 돌려준다** — 결측이
      아니라 0 이다. 그대로 두면 "이 분류는 3개월간 0% 였다"는 **거짓 문장**이 되고, 시총이
      큰 신규 종목이 끼면 분류 평균을 통째로 0 쪽으로 끌어내린다.
      실측(2026-08-31): 0.0 인 종목이 1주 0개 → 1달 8 → 3달 39 → 6달 100 으로 창 길이를
      따라 단조 증가했고, 전부 상장일이 창 시작일보다 늦었다. 8/19 시점 '단일종목' 분류의
      3개월 수익률 0.00% 가 바로 이 거짓값이었다(같은 종목들이 8/31 엔 -67%·-52%).
    ★두 번째 조건(정확히 0.0 인데 상장은 빠름)은 거래정지·데이터 공백이다(예: ACE
      러시아MSCI(합성) — 제재로 정지). 1주 이상 창에서 **정확히** 0.0000 은 실제 시세로는
      나오지 않는다. 당일 등락률에는 이 규칙을 적용하지 않는다 — 보합은 흔하다.
    """
    if v is None:
        return None
    if not _listed_before(listed, window_start):
        return None
    if v == 0.0:
        return None
    return v


def _weekdays_between(start: str | None, end: str | None) -> int:
    """(start, end] 사이의 평일 수. 기간 유입을 **일평균**으로 환산할 때 쓰는 분모다.

    ★왜 필요한가: 어제·1주·1개월 유입액을 한 축에 나란히 놓으면 1개월이 어제의 20배가
      넘어 어제 막대가 안 보이고, 게다가 모든 분류가 똑같이 "1개월>1주>어제" 모양이라
      비교할 게 없다. 일평균으로 나누면 세 막대가 같은 축에서 비교되고 "어제 유입이 평소
      대비 센가"가 바로 읽힌다.
    ★시작일은 제외한다 — FF 창의 시작일은 기준선이고 합계는 그 다음 거래일부터다
      (실측: 1w 창 08/24~08/31 의 daily 리포트가 08/25·26·27·28·31 다섯 장).
    ⚠️공휴일은 반영하지 않는다(주말만 뺀다). 한국 증시 휴장일이 연 10~15일이라 한 달 창에서
      한두 날 어긋난다 — 막대 길이의 몇 %라 순위를 뒤집지 않지만, 화면이 근사임을 밝힌다.
    """
    if not start or not end:
        return 0
    try:
        a = date.fromisoformat(start)
        b = date.fromisoformat(end)
    except ValueError:
        return 0
    n = 0
    d = a + timedelta(days=1)
    while d <= b:
        if d.weekday() < 5:
            n += 1
        d += timedelta(days=1)
    return n


def _chain(outer: float | None, inner: float | None) -> float | None:
    """누적 수익률 둘 → 안쪽을 걷어낸 구간 수익률. (1+outer)/(1+inner)-1."""
    if outer is None:
        return None
    if inner is None:
        return outer
    if inner <= -1:
        return None      # -100% 는 나눌 수 없다. 값을 지어내지 않는다.
    return (1.0 + outer) / (1.0 + inner) - 1.0


def _diff(outer: float | None, inner: float | None) -> float | None:
    """누적 순매수 둘 → 구간 순매수. 합계라 그냥 뺀다."""
    if outer is None:
        return None
    return outer - (inner or 0.0)


# 묶음 크기 때문에 자릿수를 자른다 — 861종목 × 지표 18개면 `-0.20309999999999997` 같은
# 부동소수 꼬리가 payload 의 3할이다. 억은 소수 3자리(=1천원), 수익률은 6자리(=0.0001%)로
# 자른다. 화면 표기는 소수 1~2자리라 눈에 보이는 값은 하나도 안 바뀐다.
def _r(v: float | None, nd: int) -> float | None:
    return None if v is None else round(v, nd)


def _rd(d: dict, nd: int) -> dict:
    return {k: _r(v, nd) for k, v in d.items()}


# ── 워크북 판독 (mtime+size 캐시) ────────────────────────────────────────────

# ★캐시 키에 **경로**가 들어가야 한다. 원래는 (mtime, size) 만 봤는데, `seed_archive` 가
#   같은 함수로 백업본 여러 장을 훑으면서 캐시를 갈아 치운다. 그 사이 들어온 요청이 정본을
#   달라고 했을 때 크기·시각이 우연히 겹치면 **다른 워크북의 스냅샷**을 정본으로 내놓게 된다.
#   확률은 낮지만 그렇게 틀리면 화면 전체가 조용히 다른 날을 말한다.
_CACHE: dict = {"key": None, "snap": None}


def _read_snapshot(path: str = SRC_PATH) -> dict:
    """워크북 한 장 → {asof, windows, etfs[], source_modified}.

    헤더(5행)를 COLS 의 기대 문자열과 대조한다. 어긋나면 그 열만 비우는 게 아니라
    **전체를 거부**한다 — 열이 한 칸 밀린 채 계산하면 화면이 조용히 거짓말을 한다.
    """
    st = os.stat(path)
    key = (os.path.abspath(path), st.st_mtime_ns, st.st_size)
    if _CACHE["key"] == key and _CACHE["snap"] is not None:
        return _CACHE["snap"]

    with open(path, "rb") as f:
        blob = f.read()

    import openpyxl

    wb = openpyxl.load_workbook(
        io.BytesIO(blob), data_only=True, read_only=True, keep_links=False
    )
    try:
        ws = wb[SHEET]
        head = [row for row in ws.iter_rows(min_row=1, max_row=HEADER_ROW, values_only=True)]

        asof = _asof_date(head[ASOF_CELL[0] - 1][ASOF_CELL[1] - 1])
        hdr = head[HEADER_ROW - 1]
        for key, (idx, want) in COLS.items():
            got = _txt(hdr[idx] if idx < len(hdr) else None)
            if got != want:
                raise ValueError(f"열 배치가 다릅니다: {idx + 1}번째 열이 {got!r} (기대 {want!r})")
        # 선택 열은 있으면 읽고 없으면 건너뛴다(헤더가 맞을 때만 읽는다 —
        # 자리만 보고 읽으면 나중에 다른 열이 그 자리에 오면 조용히 틀린다).
        opt = {
            key: idx
            for key, (idx, want) in OPTIONAL_COLS.items()
            if _txt(hdr[idx] if idx < len(hdr) else None) == want
        }

        # 1·2행 15~22열 = MMT 4창(15~18) + FF 4창(19~22)의 시작일/끝일.
        # 지금 워크북은 두 벌이 같은 날짜라 한 벌만 화면에 쓴다. 그래도 대조는 한다 —
        # 언젠가 갈라지면 "수익률 창"으로 "자금 창"을 설명하는 말이 되기 때문이다.
        def _win(base: int) -> dict:
            return {
                h: {
                    "start": _asof_date(head[0][base + i]) if len(head[0]) > base + i else None,
                    "end": _asof_date(head[1][base + i]) if len(head[1]) > base + i else None,
                }
                for i, h in enumerate(_HORIZONS)
            }

        windows = _win(14)
        ff_windows = _win(18)
        window_mismatch = [h for h in _HORIZONS if windows[h] != ff_windows[h]]

        etfs = []
        for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            code = _txt(r[COLS["code"][0]]) if len(r) > COLS["code"][0] else ""
            name = _txt(r[COLS["name"][0]]) if len(r) > COLS["name"][0] else ""
            if not code or not name:
                continue
            rec = {"code": code, "name": name}
            for key, (idx, _w) in COLS.items():
                if key in ("code", "name"):
                    continue
                v = r[idx] if idx < len(r) else None
                if key in ("listed", "country", "gubun", "big", "mid", "small"):
                    rec[key] = _txt(v)
                else:
                    rec[key] = _num(v)
            for key, idx in OPTIONAL_DEFAULT.items():
                rec[key] = idx
            if "interest" in opt:
                rec["interest"] = bool(_num(r[opt["interest"]] if opt["interest"] < len(r) else None))
            etfs.append(rec)
    finally:
        wb.close()

    # ★여기서 한 번 걸러 두면 화면·적재·이력이 전부 같은 값을 본다. 소비처마다 거르면
    #   한 곳을 고칠 때 다른 곳이 옛 값을 계속 말한다.
    masked = 0
    for e in etfs:
        for h in _HORIZONS:
            k = "mmt_" + h
            before = e[k]
            e[k] = _valid_return(before, e["listed"], windows[h]["start"])
            if before is not None and e[k] is None:
                masked += 1

    snap = {
        "asof": asof,
        "windows": windows,
        "ff_windows": ff_windows,
        "window_mismatch": window_mismatch,
        "masked_returns": masked,
        "etfs": etfs,
        "source_modified": datetime.fromtimestamp(st.st_mtime, _KST).strftime("%Y-%m-%d %H:%M"),
    }
    _CACHE["key"] = key
    _CACHE["snap"] = snap
    return snap


# ── ETF 한 줄 → 화면이 쓰는 지표 ─────────────────────────────────────────────

def _etf_metrics(e: dict) -> dict:
    """누적·구간을 ETF 단위에서 확정한다.

    ★수익률 구간은 여기서만 만든다. 분류 평균끼리 체인하면(집계 후 체인) 분류의 종목
      구성이 창마다 달라 값이 미끄러진다 — 종목 단위 체인이 유일하게 옳다.
    """
    net_cum = {"d": e["net"], "1w": e["ff_1w"], "1m": e["ff_1m"],
               "3m": e["ff_3m"], "6m": e["ff_6m"]}
    ret_cum = {"d": e["chg"], "1w": e["mmt_1w"], "1m": e["mmt_1m"],
               "3m": e["mmt_3m"], "6m": e["mmt_6m"]}
    net_iv, ret_iv = {}, {}
    for spec in INTERVALS:
        o, i = spec["outer"], spec["inner"]
        net_iv[spec["key"]] = _diff(net_cum[o], net_cum[i] if i else None)
        ret_iv[spec["key"]] = _chain(ret_cum[o], ret_cum[i] if i else None)
    return {"net_cum": net_cum, "ret_cum": ret_cum, "net_iv": net_iv, "ret_iv": ret_iv}


def _blank_bucket(label: str, path: list[str]) -> dict:
    return {
        "key": label, "label": label, "path": path, "n": 0,
        "mcap": 0.0, "amt": 0.0,
        "net_cum": dict.fromkeys(_PERIOD_KEYS, 0.0),
        "net_iv": dict.fromkeys(_IV_KEYS, 0.0),
        # 수익률은 (Σ w·r, Σ w) 를 따로 모은다 — 결측 종목의 시총이 분모에 남으면
        # 그 분류만 0 쪽으로 끌려 내려간다.
        "_rw": dict.fromkeys(_PERIOD_KEYS, 0.0),
        "_rwd": dict.fromkeys(_PERIOD_KEYS, 0.0),
        "_rs": dict.fromkeys(_PERIOD_KEYS, 0.0),
        "_rn": dict.fromkeys(_PERIOD_KEYS, 0),
        "_iw": dict.fromkeys(_IV_KEYS, 0.0),
        "_iwd": dict.fromkeys(_IV_KEYS, 0.0),
        "_is": dict.fromkeys(_IV_KEYS, 0.0),
        "_in": dict.fromkeys(_IV_KEYS, 0),
    }


def _accumulate(b: dict, e: dict, m: dict) -> None:
    b["n"] += 1
    mcap = e["mcap"] or 0.0
    b["mcap"] += mcap
    b["amt"] += e["amt"] or 0.0
    for k in _PERIOD_KEYS:
        v = m["net_cum"][k]
        if v is not None:
            b["net_cum"][k] += v
        r = m["ret_cum"][k]
        if r is not None:
            b["_rw"][k] += r * mcap
            b["_rwd"][k] += mcap
            b["_rs"][k] += r
            b["_rn"][k] += 1
    for k in _IV_KEYS:
        v = m["net_iv"][k]
        if v is not None:
            b["net_iv"][k] += v
        r = m["ret_iv"][k]
        if r is not None:
            b["_iw"][k] += r * mcap
            b["_iwd"][k] += mcap
            b["_is"][k] += r
            b["_in"][k] += 1


def _finalize(b: dict) -> dict:
    """가중/단순 평균을 확정하고 내부 누산기를 걷어낸다."""
    out = {k: v for k, v in b.items() if not k.startswith("_")}
    out["ret_cum"] = {k: (b["_rw"][k] / b["_rwd"][k] if b["_rwd"][k] else None)
                      for k in _PERIOD_KEYS}
    out["ret_cum_eq"] = {k: (b["_rs"][k] / b["_rn"][k] if b["_rn"][k] else None)
                         for k in _PERIOD_KEYS}
    out["ret_iv"] = {k: (b["_iw"][k] / b["_iwd"][k] if b["_iwd"][k] else None)
                     for k in _IV_KEYS}
    out["ret_iv_eq"] = {k: (b["_is"][k] / b["_in"][k] if b["_in"][k] else None)
                        for k in _IV_KEYS}
    # 강도 — 절대 억원은 규모가 이긴다. 시총 대비 몇 %가 들어왔나로 같이 본다.
    mc = b["mcap"] or 0.0
    out["ratio_cum"] = {k: (b["net_cum"][k] / mc * 100.0 if mc else None)
                        for k in _PERIOD_KEYS}
    out["ratio_iv"] = {k: (b["net_iv"][k] / mc * 100.0 if mc else None)
                       for k in _IV_KEYS}
    out["mcap"] = _r(out["mcap"], 1)
    out["amt"] = _r(out["amt"], 1)
    for k in ("net_cum", "net_iv"):
        out[k] = _rd(out[k], 3)
    for k in ("ret_cum", "ret_cum_eq", "ret_iv", "ret_iv_eq", "ratio_cum", "ratio_iv"):
        out[k] = _rd(out[k], 6)
    return out


def etf_group_key(e: dict, axis_key: str) -> str:
    """ETF 한 줄이 그 축에서 속하는 분류 키. `_group` 과 **같은 식**이어야 한다.

    ★화면이 이 계산을 다시 하지 않도록 payload 에 실어 보낸다. 예전엔 프런트가 같은 규칙을
      한 벌 더 갖고 있었는데, 표기 접기(`_fold`)가 들어오면서 두 곳이 갈릴 수 있게 됐다 —
      갈리면 분류를 눌러도 상세 표가 비고, 그게 "데이터가 없다"로 보여 오진을 부른다.
    """
    axis = _AXIS_BY_KEY[axis_key]
    parts = [e.get(c) or "미분류" for c in axis["path"]] + [e.get(axis["col"]) or "미분류"]
    return " / ".join(parts)


def _group(etfs: list[dict], metrics: list[dict], axis_key: str) -> list[dict]:
    """한 축으로 묶는다. 라벨이 빈 종목은 '미분류' 로 모은다(조용히 버리지 않는다)."""
    axis = _AXIS_BY_KEY[axis_key]
    col, path_cols = axis["col"], axis["path"]
    buckets: dict[tuple, dict] = {}
    for e, m in zip(etfs, metrics):
        label = e.get(col) or "미분류"
        path = [e.get(c) or "미분류" for c in path_cols]
        # 묶는 건 접은 키로, 보이는 건 처음 만난 철자로. (`Top10`/`TOP10` 이 한 워크북
        # 안에서도 둘 다 나온다 — 접지 않으면 같은 분류가 두 행이 된다.)
        k = tuple(_fold(x) for x in (*path, label))
        b = buckets.get(k)
        if b is None:
            b = buckets[k] = _blank_bucket(label, path)
            b["key"] = " / ".join((*path, label))
        # 대표 철자를 쓴 키를 ETF 에 되돌려 준다 — 접힌 쪽(`TOP10`)도 대표(`Top10`)를 가리킨다.
        e.setdefault("_gkeys", {})[axis_key] = b["key"]
        _accumulate(b, e, m)
    rows = [_finalize(b) for b in buckets.values()]
    rows.sort(key=lambda r: -(r["net_cum"]["3m"] or 0.0))
    return rows


# ── payload ──────────────────────────────────────────────────────────────────

def build_snapshot() -> dict:
    """오늘의 분류별 자금·수익률 한 장. 전 축·전 기간을 한 번에 실어 보낸다.

    ★화면이 축·기간을 바꿀 때 재요청하지 않는다 — 축 5개·기간 9개를 다 담아도 묶음이
      크지 않고(수백 KB), 집계식이 서버 한 곳에만 있어 두 곳으로 갈라지지 않는다.
    """
    out: dict = {
        "generated_at": datetime.now(_KST).strftime("%Y-%m-%d %H:%M:%S"),
        "asof": None,
        "source_modified": None,
        "note": None,
        "axes": [{"key": a["key"], "label": a["label"]} for a in AXES],
        "periods": [dict(p) for p in PERIODS],
        "intervals": [dict(s) for s in INTERVALS],
        "windows": {},
        "groups": {},
        "etfs": [],
        "totals": None,
        "history_days": 0,
        "masked_returns": 0,
    }
    try:
        snap = _read_snapshot()
    except FileNotFoundError:
        out["note"] = f"원천 워크북이 없습니다 — {SRC_PATH}"
        return out
    except Exception as exc:  # 열 배치 어긋남·손상 등
        out["note"] = f"워크북을 읽지 못했습니다 — {exc}"
        return out

    etfs = snap["etfs"]
    metrics = [_etf_metrics(e) for e in etfs]
    out["asof"] = snap["asof"]
    out["source_modified"] = snap["source_modified"]
    out["windows"] = snap["windows"]
    out["masked_returns"] = snap.get("masked_returns", 0)
    for p in out["periods"]:
        w = snap["windows"].get(p["span"]) if p["span"] else None
        p["start"] = (w or {}).get("start") or snap["asof"]
        p["end"] = (w or {}).get("end") or snap["asof"]
        # 당일은 거래일 1 일. 나머지는 창 안의 평일 수(주말만 제외).
        p["days"] = 1 if p["span"] is None else max(_weekdays_between(p["start"], p["end"]), 1)
    for s in out["intervals"]:
        inner = snap["windows"].get(s["inner"]) if s["inner"] else None
        outer = snap["windows"].get(s["outer"]) or {}
        s["start"] = outer.get("start")
        s["end"] = (inner or {}).get("start") or outer.get("end")
        s["days"] = max(_weekdays_between(s["start"], s["end"]), 1)

    if snap.get("window_mismatch"):
        out["note"] = (
            "수익률(MMT)과 순매수(FF)의 기간 창이 다릅니다 — "
            + ", ".join(snap["window_mismatch"])
            + ". 화면의 기간 표기는 수익률 창 기준입니다."
        )

    for a in AXES:
        out["groups"][a["key"]] = _group(etfs, metrics, a["key"])

    out["etfs"] = [
        {
            "code": e["code"], "name": e["name"],
            # 상장일 — 화면의 '신규 상장' 칸이 쓴다. 셀이 날짜형이면 시:분이 붙어 오므로
            # 앞 10자만 남긴다(YYYY-MM-DD).
            "listed": (e["listed"] or "")[:10],
            "country": e["country"], "gubun": e["gubun"],
            "big": e["big"], "mid": e["mid"], "small": e["small"],
            "mcap": _r(e["mcap"], 1), "amt": _r(e["amt"], 1), "price": e["price"],
            "interest": e["interest"],
            "net_cum": _rd(m["net_cum"], 3), "net_iv": _rd(m["net_iv"], 3),
            "ret_cum": _rd(m["ret_cum"], 6), "ret_iv": _rd(m["ret_iv"], 6),
            # AXES 순서대로의 분류 키. 화면은 이걸로 조인한다(규칙을 두 곳에 두지 않는다).
            "gkeys": [(e.get("_gkeys") or {}).get(a["key"], "") for a in AXES],
        }
        for e, m in zip(etfs, metrics)
    ]

    total = _blank_bucket("전체", [])
    for e, m in zip(etfs, metrics):
        _accumulate(total, e, m)
    out["totals"] = _finalize(total)

    # 적재는 판독 뒤 조용히 한 번. 실패해도 화면은 그대로 나간다.
    try:
        out["history_days"] = ingest(snap)
    except Exception:
        out["history_days"] = 0
    return out


# ── 일별 누적 (read-through 적재) ────────────────────────────────────────────

# ★적재된 값은 **판독 규칙까지 굳은 결과물**이다(예: MMT 무효값 마스킹). 규칙을 고치면
#   이미 쌓인 행은 옛 규칙이라 화면이 시점마다 다른 말을 한다. 그래서 버전을 두고,
#   달라지면 통째로 버리고 다시 굽는다 — 원천(워크북·백업본)이 남아 있어 복구가 싸다.
SCHEMA_VERSION = "2"

_DDL = """
CREATE TABLE IF NOT EXISTS meta (k TEXT PRIMARY KEY, v TEXT);
CREATE TABLE IF NOT EXISTS snapshot (
  asof     TEXT NOT NULL,
  code     TEXT NOT NULL,
  name     TEXT,
  country  TEXT, gubun TEXT, big TEXT, mid TEXT, small TEXT,
  amt REAL, net REAL, chg REAL, price REAL, mcap REAL,
  mmt_1w REAL, mmt_1m REAL, mmt_3m REAL, mmt_6m REAL,
  ff_1w REAL, ff_1m REAL, ff_3m REAL, ff_6m REAL,
  PRIMARY KEY (asof, code)
);
CREATE INDEX IF NOT EXISTS ix_snapshot_asof ON snapshot(asof);
"""

_INSERT = (
    "INSERT OR REPLACE INTO snapshot (asof,code,name,country,gubun,big,mid,small,"
    "amt,net,chg,price,mcap,mmt_1w,mmt_1m,mmt_3m,mmt_6m,ff_1w,ff_1m,ff_3m,ff_6m) "
    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
)


def _connect() -> sqlite3.Connection:
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    con = sqlite3.connect(DB_PATH, timeout=10)
    con.executescript(_DDL)
    row = con.execute("SELECT v FROM meta WHERE k='schema'").fetchone()
    if row is None or row[0] != SCHEMA_VERSION:
        con.execute("DELETE FROM snapshot")
        con.execute("INSERT OR REPLACE INTO meta (k,v) VALUES ('schema',?)", (SCHEMA_VERSION,))
        con.commit()
    return con


def _rows_of(snap: dict) -> list[tuple]:
    asof = snap["asof"]
    return [
        (asof, e["code"], e["name"], e["country"], e["gubun"], e["big"], e["mid"],
         e["small"], e["amt"], e["net"], e["chg"], e["price"], e["mcap"],
         e["mmt_1w"], e["mmt_1m"], e["mmt_3m"], e["mmt_6m"],
         e["ff_1w"], e["ff_1m"], e["ff_3m"], e["ff_6m"])
        for e in snap["etfs"]
    ]


def ingest(snap: dict | None = None) -> int:
    """스냅샷을 적재하고 보유 일수를 돌려준다. 같은 기준일은 다시 넣지 않는다.

    워크북은 매일 **덮어쓰기**라 이 적재가 유일한 과거 보관이다. 스케줄러를 두지 않고
    판독 때마다 확인하는 read-through 로 만든 이유: 화면이 열려 있는 한 스스로 낫는다
    (기동 순서·잡 실패에 기대지 않는다).
    """
    if snap is None:
        snap = _read_snapshot()
    asof = snap.get("asof")
    if not asof:
        return 0
    con = _connect()
    try:
        if con.execute("SELECT 1 FROM snapshot WHERE asof=? LIMIT 1", (asof,)).fetchone() is None:
            con.executemany(_INSERT, _rows_of(snap))
            con.commit()
        (days,) = con.execute("SELECT COUNT(DISTINCT asof) FROM snapshot").fetchone()
        return int(days or 0)
    finally:
        con.close()



# ── 과거 스냅샷 복원 ─────────────────────────────────────────────────────────
# 원천 폴더에는 운용역이 남긴 백업본이 몇 장 있다. 스키마가 같은 것만 골라 한 번 적재하면
# 이력이 오늘 하루에서 몇 달로 늘어난다. ★성기다는 사실 자체를 화면이 말해야 하므로 여기서
# 보간하지 않는다 — 그날 실제로 있었던 관측만 넣는다.
# ⚠️백업 폴더에는 SpaceX자금유입·중국AI 같은 남의 파일도 있어 확장자만으로는 못 고른다.
ARCHIVE_DIRS = ["", "Backup", "회의자료백업"]
ARCHIVE_NAME_HINT = "모니터링"


def _archive_paths() -> list[str]:
    root = os.path.dirname(SRC_PATH)
    out: list[str] = []
    for sub in ARCHIVE_DIRS:
        d = os.path.join(root, sub) if sub else root
        try:
            names = sorted(os.listdir(d))
        except OSError:
            continue
        for n in names:
            if not n.lower().endswith((".xlsm", ".xlsx")):
                continue
            if ARCHIVE_NAME_HINT not in n or n.startswith("~$"):
                continue
            out.append(os.path.join(d, n))
    return out


def seed_archive() -> dict:
    """백업 워크북 중 아직 없는 기준일을 적재하고 요약을 돌려준다.

    열 배치가 다른 옛 사본(_v2·_vvv)은 _read_snapshot 이 거부하므로 여기서 따로 거를 게
    없다 — **거부된 이유를 세어서 알린다**(조용히 빠지면 왜 이력이 짧은지 알 수 없다).
    """
    con = _connect()
    try:
        have = {r[0] for r in con.execute("SELECT DISTINCT asof FROM snapshot")}
    finally:
        con.close()

    added, skipped = [], []
    for path in _archive_paths():
        _CACHE["key"] = None          # 캐시는 정본 한 장을 위한 것 — 훑을 땐 비운다
        try:
            snap = _read_snapshot(path)
        except Exception as exc:  # noqa: BLE001
            skipped.append(f"{os.path.basename(path)}: {exc}")
            continue
        if not snap["asof"] or snap["asof"] in have:
            continue
        ingest(snap)
        have.add(snap["asof"])
        added.append(snap["asof"])
    _CACHE["key"] = None              # 정본이 다시 읽히도록 캐시를 비운 채 끝낸다
    con = _connect()
    try:
        (days,) = con.execute("SELECT COUNT(DISTINCT asof) FROM snapshot").fetchone()
    finally:
        con.close()
    return {"added": sorted(added), "skipped": skipped, "days": int(days or 0)}


# ── 시점별 추이 ──────────────────────────────────────────────────────────────
# ★★2026-09-01 개편. 처음엔 "일별 누적 순매수"였는데, 그건 **날마다 관측이 있어야만** 말이
#   되는 그림이다. 복원한 과거는 성긴 스냅샷(4·5·6·8월)이라 그 위에 누적선을 그으면 빠진
#   날의 자금이 0 인 것처럼 보인다 — 없는 사실을 그리는 셈이다.
#   그래서 축을 바꿨다: **각 시점에서 본 그 기간의 값**을 그린다. 5/15 의 3개월 누적,
#   6/8 의 3개월 누적, … 처럼 관측 하나가 이미 누적이라 성겨도 정확하다. 일별 스냅샷이
#   쌓일수록 같은 그림이 저절로 촘촘해진다.
HISTORY_METRICS = [
    {"key": "net", "label": "개인 순매수", "unit": "억"},
    {"key": "ret", "label": "수익률", "unit": "%"},
    {"key": "mcap", "label": "시총", "unit": "억"},
]
# (지표, 기간) → sqlite 열. 시총은 기간이 없다(그 시점의 값 하나뿐).
_HIST_COL = {
    ("net", "d"): "net", ("net", "1w"): "ff_1w", ("net", "1m"): "ff_1m",
    ("net", "3m"): "ff_3m", ("net", "6m"): "ff_6m",
    ("ret", "d"): "chg", ("ret", "1w"): "mmt_1w", ("ret", "1m"): "mmt_1m",
    ("ret", "3m"): "mmt_3m", ("ret", "6m"): "mmt_6m",
}


def build_history(
    axis: str = DEFAULT_AXIS,
    metric: str = "net",
    period: str = "3m",
    days: int = 400,
) -> dict:
    """시점별 추이 — 적재된 스냅샷마다 분류별 값을 하나씩.

    ★수익률은 **그 시점 스냅샷 자신의 시총**으로 가중한다. 표(build_snapshot)는 한 시점만
      다뤄 오늘 시총을 쓸 수밖에 없지만, 여기는 과거 시총이 실제로 있으므로 근사가 아니다.
    """
    axis_key = axis if axis in _AXIS_BY_KEY else DEFAULT_AXIS
    col_group = _AXIS_BY_KEY[axis_key]["col"]
    metric = metric if any(m["key"] == metric for m in HISTORY_METRICS) else "net"
    period = period if period in _PERIOD_KEYS else "3m"
    out = {
        "generated_at": datetime.now(_KST).strftime("%Y-%m-%d %H:%M:%S"),
        "axis": axis_key,
        "metric": metric,
        "period": period,
        "metrics": [dict(m) for m in HISTORY_METRICS],
        "dates": [],
        "series": [],
        "note": None,
    }
    try:
        con = _connect()
    except OSError as exc:
        out["note"] = f"이력 저장소를 열지 못했습니다 — {exc}"
        return out

    since = (date.today() - timedelta(days=days)).isoformat()
    grp = f"COALESCE(NULLIF({col_group},''),'미분류')"
    vcol = "mcap" if metric == "mcap" else _HIST_COL[(metric, period)]
    if metric == "ret":
        # 결측 종목의 시총은 분모에서 빼야 그 분류만 0 쪽으로 끌려가지 않는다.
        sql = (
            f"SELECT asof, {grp} AS g, "
            f"SUM(COALESCE({vcol},0)*COALESCE(mcap,0)), "
            f"SUM(CASE WHEN {vcol} IS NULL THEN 0 ELSE COALESCE(mcap,0) END) "
            "FROM snapshot WHERE asof >= ? GROUP BY asof, g ORDER BY asof"
        )
    else:
        sql = (
            f"SELECT asof, {grp} AS g, SUM(COALESCE({vcol},0)), NULL "
            "FROM snapshot WHERE asof >= ? GROUP BY asof, g ORDER BY asof"
        )
    try:
        rows = con.execute(sql, (since,)).fetchall()
    finally:
        con.close()

    if not rows:
        out["note"] = "아직 쌓인 스냅샷이 없습니다."
        return out

    # ★묶기는 **접은 키**로 한다. SQL 의 GROUP BY 로는 `MSCI KOREA`/`MSCI Korea` 가 두 줄이
    #   되어 한 분류의 시계열이 4월에서 끊기고 8월에 새로 시작한 것처럼 보인다.
    #   표시 철자는 **가장 최근 시점의 것**을 쓴다(지금 회의에서 쓰는 말).
    dates = sorted({r[0] for r in rows})
    idx = {d: i for i, d in enumerate(dates)}
    acc: dict[str, dict] = {}
    for asof, g, a, b in rows:
        k = _fold(g)
        e = acc.get(k)
        if e is None:
            e = acc[k] = {"label": g, "num": [0.0] * len(dates),
                          "den": [0.0] * len(dates), "seen": [False] * len(dates)}
        i = idx[asof]
        e["label"] = g          # rows 가 asof 오름차순이라 마지막이 최신 철자다
        e["num"][i] += a or 0.0
        e["den"][i] += (b or 0.0) if metric == "ret" else 0.0
        e["seen"][i] = True

    series = []
    for k, e in acc.items():
        vals: list[float | None] = []
        for i in range(len(dates)):
            if not e["seen"][i]:
                vals.append(None)
            elif metric == "ret":
                vals.append(_r(e["num"][i] / e["den"][i], 6) if e["den"][i] else None)
            else:
                vals.append(_r(e["num"][i], 3))
        series.append({
            "key": k,
            "label": e["label"],
            "values": vals,
            # 정렬 기준 = 가장 최근 관측. 마지막 값이 곧 지금 상태다.
            "last": next((v for v in reversed(vals) if v is not None), None),
        })
    series.sort(key=lambda x: -(x["last"] or 0))
    out["dates"] = dates
    out["series"] = series
    return out
