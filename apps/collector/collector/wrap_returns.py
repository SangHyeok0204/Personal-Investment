"""WRAP ① 전전일→전일 close-to-close 수익률 — Price 시트 판독 (2026-07-22).

``이상 포트폴리오 수익률.xlsx`` 의 ``Price`` 시트에서 종목별 최근 2영업일 종가와
날짜별 환율을 뽑아, 포트폴리오의 종가-대-종가 수익률(USD / 원화)을 계산할 수 있게
한다. 시트 구조 (사용자 확인 D1:BV95):

  * 주가 블록 = **D:BV** — D열=날짜, 2행=티커(E..), 3행~=일별 종가. A:B 환율 블록과
    행 정렬이 다르므로 각각 자기 날짜열로 읽는다.
  * 환율 블록 = **A:B** — A열=날짜, B열=USD/KRW. (주식보다 날짜가 더 많음 → 조회 안전)

순수 판독기 — 파일 기록 없음(read-only invariant), 예외 시 None.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

# 주가 블록: D(4)=날짜, E(5)~ = 티커별 종가. 환율 블록: A(1)=날짜, B(2)=USD/KRW.
_DATE_COL = 4  # D (1-indexed)
_FIRST_TICKER_COL = 5  # E
_TICKER_HEADER_ROW = 2
_MAX_COL = 74  # BV
_MAX_ROW = 400


def _num(v) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _as_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def expected_basis_date(today: date) -> date:
    """오늘(KST) 기준 '가장 최근 확정된 미장 종가일' = 오늘 직전 평일.

    미장 7/21 종가는 KST 7/22 05:00 확정 → 7/22 어느 시각이든 기대 최신종가일=7/21.
    주말은 건너뛴다(미 공휴일은 미반영 — 드물게 '미갱신' 오탐 가능, 날짜라벨은 정확).
    """
    d = today - timedelta(days=1)
    while d.weekday() >= 5:  # 5=토, 6=일
        d -= timedelta(days=1)
    return d


def read_price_closes(src_path: Path, sheet: str = "Price") -> dict | None:
    """Price 시트 → 티커별 (전전일 종가, 전일 종가) + 날짜 + 환율.

    반환 dict:
      closes:   {TICKER: (prev_close, last_close)}  (둘 다 유효한 종목만)
      last_date/prev_date: date
      fx_last/fx_prev:     float | None  (해당 날짜의 USD/KRW)
    실패 시 None.
    """
    try:
        wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        names = wb.sheetnames
        target = (
            sheet
            if sheet in names
            else next((s for s in names if s.lower() == sheet.lower()), None)
        )
        if not target:
            return None
        ws = wb[target]
        rows = list(
            ws.iter_rows(min_row=1, max_row=_MAX_ROW, max_col=_MAX_COL, values_only=True)
        )
    except Exception:
        return None
    finally:
        wb.close()

    if len(rows) < 3:
        return None

    # 티커 헤더 (2행, E열~)
    header = rows[_TICKER_HEADER_ROW - 1]
    tickers: dict[int, str] = {}  # 0-indexed col -> ticker
    for ci in range(_FIRST_TICKER_COL - 1, len(header)):
        t = header[ci]
        if isinstance(t, str) and t.strip():
            tickers[ci] = t.strip().upper()
    if not tickers:
        return None

    # 주가 블록: D열 날짜 + E열 값이 모두 있는 행만 (3행~)
    di = _DATE_COL - 1
    ei = _FIRST_TICKER_COL - 1
    stock_rows = [
        rv
        for rv in rows[2:]
        if len(rv) > ei and rv[di] is not None and rv[ei] is not None
    ]
    if len(stock_rows) < 2:
        return None
    last_rv, prev_rv = stock_rows[-1], stock_rows[-2]
    last_date, prev_date = _as_date(last_rv[di]), _as_date(prev_rv[di])
    if last_date is None or prev_date is None:
        return None

    closes: dict[str, tuple[float, float]] = {}
    for ci, tk in tickers.items():
        lp = _num(last_rv[ci]) if ci < len(last_rv) else None
        pp = _num(prev_rv[ci]) if ci < len(prev_rv) else None
        if lp and pp and lp > 0 and pp > 0:
            closes[tk] = (pp, lp)

    # 환율 블록: A열 날짜 -> B열 환율
    fxmap: dict[date, float] = {}
    for rv in rows[2:]:
        da = _as_date(rv[0]) if rv else None
        fb = _num(rv[1]) if len(rv) > 1 else None
        if da is not None and fb:
            fxmap[da] = fb

    return {
        "closes": closes,
        "last_date": last_date,
        "prev_date": prev_date,
        "fx_last": fxmap.get(last_date),
        "fx_prev": fxmap.get(prev_date),
    }


def compute_ret1(
    rows_by_pf: dict[str, list[dict]],
    price: dict,
    cash_tickers,
) -> dict[str, dict]:
    """포트폴리오별 ① 종가-대-종가 수익률 (USD / 원화).

    usd = Σ 비중 × (전일종가/전전일종가 − 1)
    krw = Σ 비중 × (전일종가·환율_전일 / 전전일종가·환율_전전일 − 1)
    Price 시트 티커는 전부 미국 상장 → 종목이 closes 에 있으면 환율 적용.
    """
    closes = price["closes"]
    fx_last, fx_prev = price.get("fx_last"), price.get("fx_prev")
    use_fx = bool(fx_last and fx_prev)
    basis = price["last_date"].strftime("%Y%m%d")
    prev = price["prev_date"].strftime("%Y%m%d")

    out: dict[str, dict] = {}
    for key, rows in rows_by_pf.items():
        usd_sum = krw_sum = 0.0
        for h in rows:
            t = str(h["ticker"]).upper()
            if t in cash_tickers:
                continue
            pair = closes.get(t)
            if not pair:
                continue
            pp, lp = pair
            r_usd = (lp / pp - 1.0) * 100.0
            r_krw = (
                ((lp * fx_last) / (pp * fx_prev) - 1.0) * 100.0 if use_fx else r_usd
            )
            w = (h["weight_pct"] or 0.0) / 100.0
            usd_sum += w * r_usd
            krw_sum += w * r_krw
        out[key] = {"usd": usd_sum, "krw": krw_sum, "basis": basis, "prev": prev}
    return out
