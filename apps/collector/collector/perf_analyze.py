"""[성과보고] 엑셀 실시간 분석 엔진 (2026-07-28).

대시보드 [분석 시작] 버튼이 부르는 계산기. 정기미팅 폴더(:ro)의 운용역 소스 엑셀을
직접 읽어, 성과보고와 **동일한 형태**의 정량 분석 블록을 매번 같은 순서로 만든다.

    일간 → 운용펀드 데일리 성과보고.xlsx      (단일 비중 × 당일 종목수익률)
    주간 → 운용펀드 데일리 성과보고(1W).xlsx  (기간별 비중 × 일별 가격 → 복리 체인링크)

계산 방법은 performance-brief 스킬의 검증된 엔진(scripts/compute.py)을 그대로 옮긴 것이다
— 시트 A2/A5 의 SUMPRODUCT 표시값이 아니라 여기서 재계산한 값을 쓰고, 차이가 나는
사유(포트 변경 타이밍 미반영 등)는 warnings 로 올린다.

만들지 않는 것: 시장 스트립·스토리 카드·관전 포인트. 뉴스 조사와 서사가 필요해
엑셀만으로는 나오지 않는다(스킬/LLM 몫).
"""
from __future__ import annotations

import json
import os
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

PERF_BRIEF_ROOT = os.environ.get("PERF_BRIEF_DIR", "/srv/legacy/perf_brief")

SOURCES = {
    "daily": "운용펀드 데일리 성과보고.xlsx",
    "weekly": "운용펀드 데일리 성과보고(1W).xlsx",
}
WRAP_SHEETS = ["한투미국AI코어테크", "토러스글로벌성장랩"]  # [당사, BM]
FUND_SHEET = "글로벌전기차펀드"

# 티커 접미사 → 국가 (compute.py CMAP 동일)
CMAP = {"US": "미국", "KS": "한국", "C2": "중국A", "HK": "홍콩", "SW": "유럽",
        "GR": "유럽", "FP": "유럽", "AU": "호주", "CN": "캐나다", "JP": "일본", "TT": "대만"}

WEEKDAY_KO = ["월", "화", "수", "목", "금", "토", "일"]

# 랩 분야 분류표. 마운트에 있으면 그걸 우선 쓴다(재빌드 없이 갱신 가능).
_BUNDLED_MAP = os.path.join(os.path.dirname(__file__), "assets", "wrap_sector_map.json")
_MOUNTED_MAP = os.path.join(PERF_BRIEF_ROOT, "wrap_sector_map.json")

# 차트에 실을 상·하위 종목 수 (양수 N + 음수 N).
TOP_N = 5


# ── 유틸 ────────────────────────────────────────────────────────────────

def _num(v):
    if v is None or isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:  # noqa: BLE001
        return None


def _dkey(v) -> str:
    """블록 가격행의 날짜 키. datetime → 'YYYY-MM-DD'."""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def _ko_date(iso: str) -> str:
    y, m, d = (int(x) for x in iso.split("-"))
    return f"{y}.{m:02d}.{d:02d} ({WEEKDAY_KO[date(y, m, d).weekday()]})"


def _prev_business_day(d: date) -> date:
    """직전 평일. 데일리 엑셀에는 기준일이 없어 소스 저장일로부터 역산한다.
    (휴장일은 반영하지 못하므로 화면에는 소스 저장시각을 함께 노출한다)"""
    step = 1
    while True:
        p = d - timedelta(days=step)
        if p.weekday() < 5:
            return p
        step += 1


def _load_sector_map() -> dict:
    for path in (_MOUNTED_MAP, _BUNDLED_MAP):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
        except (OSError, ValueError):
            continue
    return {}


def _weight_label(wf: float, wb: float) -> str:
    """초과수익 분해 행 라벨 — 원본 리포트의 표기 규칙 그대로."""
    if wf <= 0 and wb > 0:
        return "미보유"
    if wb <= 0 and wf > 0:
        return "보유"
    return "오버웨이트" if wf > wb else "언더웨이트"


def _pct(v: float, digits: int = 2) -> str:
    return f"{'+' if v > 0 else ''}{v * 100:.{digits}f}%"


def _bp(v: float) -> str:
    n = round(v * 10000)
    return f"{'+' if n > 0 else ''}{n}bp"


def _top_bottom(items: list[dict], n: int = TOP_N) -> list[dict]:
    """값 기준 상위 n + 하위 n (내림차순). 0 기여는 버린다."""
    live = [x for x in items if abs(x["value"]) >= 0.5]  # 0.5bp 미만은 노이즈
    live.sort(key=lambda x: -x["value"])
    if len(live) <= n * 2:
        return live
    return live[:n] + live[-n:]


# ── 데일리: 시트별 단일 비중 × 종목수익률 ──────────────────────────────

def _read_wrap_daily(ws, sheet: str, warn: list) -> dict:
    """랩 시트 메인 테이블. B=티커 C=비중 E=종목수익률(실값)."""
    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] is None:
            continue
        w, ret = _num(r[2]), _num(r[4])
        if w is None or ret is None:
            warn.append(f"[{sheet}] {r[1]}: 비중/수익률 비정상 → 제외")
            continue
        rows[str(r[1]).strip()] = {"w": w, "r": ret}
    tw = sum(x["w"] for x in rows.values())
    if not (0.5 < tw < 1.05):
        warn.append(f"[{sheet}] 비중 합 {tw * 100:.1f}% — 확인 필요")
    port = sum(x["w"] * x["r"] for x in rows.values())
    sheet_val = _num(ws["A2"].value)
    if sheet_val is not None and abs(sheet_val - port) > 0.0002:
        warn.append(
            f"[{sheet}] 시트 표시 {sheet_val * 100:+.2f}% vs 재계산 {port * 100:+.2f}% "
            "— 재계산 값을 사용"
        )
    return {"rows": rows, "return": port, "weight_sum": tw}


def _read_fund_daily(ws, warn: list) -> dict:
    """펀드 시트. B=티커 C~E=분야1/2/3 F=비중(펀드) G=비중(BM) I=종목수익률(실값)."""
    rows = {}
    missing = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] is None:
            continue
        ret = _num(r[8])
        if ret is None:
            missing.append(str(r[1]))
            ret = 0.0
        rows[str(r[1]).strip()] = {
            "s1": r[2], "s2": r[3], "s3": r[4],
            "wF": _num(r[5]) or 0.0, "wB": _num(r[6]) or 0.0, "r": ret,
        }
    if missing:
        warn.append(f"[펀드] 수익률 결측 {missing} → 0 처리")
    f = sum(x["wF"] * x["r"] for x in rows.values())
    b = sum(x["wB"] * x["r"] for x in rows.values())
    return {"rows": rows, "return": f, "bm": b}


def _check_fund_formula(path: str, warn: list) -> None:
    """펀드 A5(BM 수익률) 수식이 F열(펀드 비중)을 참조하는 알려진 오류 검사."""
    try:
        wbf = load_workbook(path, data_only=False)
    except Exception:  # noqa: BLE001
        return
    if FUND_SHEET not in wbf.sheetnames:
        return
    a5 = wbf[FUND_SHEET]["A5"].value
    if isinstance(a5, str) and "F2:" in a5.replace(" ", ""):
        warn.append("[펀드] A5(BM 수익률) 수식이 F열(펀드 비중) 참조 — G열 기준으로 재계산함")


# ── 위클리: 기간 블록 → 복리 체인링크 (compute.py 이식) ────────────────

def _find_blocks(ws) -> list[dict]:
    """K열 '시작일' 라벨 기준 기간 블록. r=라벨행, r+1=비중, r+2=티커, r+3~=일자별 가격.
    같은 (시작,종료) 중복 블록(작업용 복사본)은 첫 유효본만 취한다."""
    blocks, seen, r = [], set(), 1
    while r <= ws.max_row:
        if ws.cell(row=r, column=11).value == "시작일":
            start, end = _dkey(ws.cell(row=r, column=12).value), _dkey(ws.cell(row=r, column=14).value)
            if (start, end) in seen:
                r += 1
                continue
            T, W, c = [], [], 12
            while ws.cell(row=r + 2, column=c).value:
                T.append(str(ws.cell(row=r + 2, column=c).value).strip())
                W.append(_num(ws.cell(row=r + 1, column=c).value))
                c += 1
            P, rr = {}, r + 3
            while rr <= ws.max_row and ws.cell(row=rr, column=11).value not in (None, "시작일"):
                P[_dkey(ws.cell(row=rr, column=11).value)] = [
                    _num(ws.cell(row=rr, column=12 + i).value) for i in range(len(T))
                ]
                rr += 1
            if T and P:
                blocks.append({"start": start, "end": end, "W": W, "T": T, "P": P})
                seen.add((start, end))
            r = rr
        else:
            r += 1
    blocks.sort(key=lambda b: b["start"])
    return blocks


def _chain(ws, sheet: str, warn: list, prev_last: dict | None = None) -> dict | None:
    """블록들을 시간순 복리 체인링크. 전 종목 가격이 그대로인 행(주말·휴장)은 제외."""
    blocks = _find_blocks(ws)
    if not blocks:
        return None
    daily, last_prices = [], dict(prev_last or {})
    for bi, b in enumerate(blocks):
        dates = sorted(b["P"].keys())
        base = {}
        for i, t in enumerate(b["T"]):
            p0 = b["P"][dates[0]][i]
            if p0 is None:
                p0 = last_prices.get(t)
                if p0 is None:
                    warn.append(f"[{sheet}] {t}: 블록{bi + 1} 기준가 결측 — 첫날 0 처리")
            base[t] = p0
        prev = base
        for d in dates[1:]:
            rets = {}
            for i, t in enumerate(b["T"]):
                p1, p0 = b["P"][d][i], prev.get(t)
                if p1 is None or not p0:
                    rets[t] = 0.0
                else:
                    rets[t] = p1 / p0 - 1
            if any(abs(v) > 1e-9 for v in rets.values()):
                daily.append({
                    "date": d,
                    "return": sum((w or 0) * rets[t] for w, t in zip(b["W"], b["T"])),
                    "rets": rets,
                    "weights": dict(zip(b["T"], b["W"])),
                })
            prev = {t: (b["P"][d][i] if b["P"][d][i] is not None else prev.get(t))
                    for i, t in enumerate(b["T"])}
        last_prices.update({t: p for t, p in prev.items() if p})
    if not daily:
        warn.append(f"[{sheet}] 기간 내 유효 거래일이 없음 — 가격 갱신 확인 필요")
        return None
    total = 1.0
    for x in daily:
        total *= 1 + x["return"]
    return {"daily": daily, "period_return": total - 1, "last_prices": last_prices,
            "start": sorted(blocks[0]["P"].keys())[0]}


# ── 블록 조립 ───────────────────────────────────────────────────────────

def _bars(title, unit, rows, caption=None, dual=False):
    return {
        "type": "dualBars" if dual else "bars",
        "title": title, "unit": unit, "valueUnit": "bp",
        "rows": rows, "caption": caption,
    }


def _contrib_rows(contrib: dict, rets: dict) -> list[dict]:
    """종목 기여도 행 — note 에 해당 종목 수익률을 붙인다."""
    items = [{
        "label": t.split()[0],
        "note": _pct(rets[t]) if t in rets and rets[t] is not None else None,
        "value": c * 10000,
    } for t, c in contrib.items()]
    return _top_bottom(items)


def _active_rows(active: dict, wF: dict, wB: dict) -> list[dict]:
    items = [{
        "label": f"{t.split()[0]} {_weight_label(wF.get(t, 0.0), wB.get(t, 0.0))}",
        "value": a * 10000,
    } for t, a in active.items()]
    return _top_bottom(items)


def _sector_rows(contrib_self: dict, contrib_bm: dict, mapper, level: int) -> list[dict]:
    """분류별 기여 — 당사/BM 을 같은 분류축에 올린다(듀얼 바)."""
    agg: dict[str, list[float]] = {}
    for src, idx in ((contrib_self, 0), (contrib_bm, 1)):
        for t, c in src.items():
            name = mapper(t, level)
            agg.setdefault(name, [0.0, 0.0])[idx] += c
    rows = [{"label": k, "value": v[0] * 10000, "value2": v[1] * 10000}
            for k, v in agg.items()]
    rows.sort(key=lambda x: -x["value"])
    return [r for r in rows if abs(r["value"]) >= 0.5 or abs(r["value2"]) >= 0.5]


def _country_rows(contrib: dict) -> list[dict]:
    agg: dict[str, float] = {}
    for t, c in contrib.items():
        agg[CMAP.get(str(t).split()[-1], "기타")] = agg.get(CMAP.get(str(t).split()[-1], "기타"), 0.0) + c
    rows = [{"label": k, "value": v * 10000} for k, v in agg.items()]
    rows.sort(key=lambda x: -x["value"])
    return [r for r in rows if abs(r["value"]) >= 0.5]


def _scores(label_self, r_self, r_bm, sub_self=None, sub_bm=None, sub_alpha=None):
    spread = r_self - r_bm
    return [
        {"label": label_self, "value": _pct(r_self), "tone": "pos" if r_self > 0 else "neg",
         "sub": sub_self},
        {"label": "BM", "value": _pct(r_bm), "tone": "pos" if r_bm > 0 else "neg", "sub": sub_bm},
        {"label": "BM 대비", "value": _bp(spread), "tone": "pos" if spread > 0 else "neg",
         "variant": "alpha", "sub": sub_alpha},
    ]


# ── 메인 ────────────────────────────────────────────────────────────────

def analyze(mode: str) -> dict:
    """mode='daily'|'weekly'. 성과보고와 같은 형태의 report payload 를 만든다."""
    if mode not in SOURCES:
        raise ValueError(f"unknown mode: {mode}")
    path = os.path.join(PERF_BRIEF_ROOT, SOURCES[mode])
    if not os.path.isfile(path):
        raise FileNotFoundError(SOURCES[mode])

    warn: list[str] = []
    smap = _load_sector_map()
    if not smap:
        warn.append("랩 분야 분류표(wrap_sector_map.json)를 찾지 못해 분류별 기여를 건너뜀")

    def sector_of(ticker: str, level: int) -> str:
        v = smap.get(ticker)
        return v[level] if isinstance(v, list) and len(v) > level else "미분류"

    mtime = datetime.fromtimestamp(os.path.getmtime(path))
    wb = load_workbook(path, data_only=True)
    _check_fund_formula(path, warn)

    sections = []
    if mode == "daily":
        as_of, meta = _analyze_daily(wb, sections, warn, smap, sector_of, mtime)
    else:
        as_of, meta = _analyze_weekly(wb, sections, warn, smap, sector_of)

    unmapped = sorted({t for t in meta.get("wrap_tickers", set())
                       if t not in smap}) if smap else []
    if unmapped:
        warn.append(f"분류표에 없는 랩 종목 {unmapped} → '미분류'로 집계. 분류표에 추가 필요")

    return {
        "schema": 1,
        "kind": mode,
        "asOf": as_of,
        "period": meta.get("period"),
        "writtenOn": date.today().isoformat(),
        "eyebrow": "DAILY QUANT ANALYSIS" if mode == "daily" else "WEEKLY QUANT ANALYSIS",
        "title": "운용자산 데일리 성과 분석" if mode == "daily" else "운용자산 위클리 성과 분석",
        "dateLine": meta["dateLine"],
        "dateNote": meta["dateNote"],
        "market": [],
        "sections": sections,
        "checkpoints": None,
        "footnote": meta["footnote"],
        "warnings": warn,
        "source": SOURCES[mode],
        "sourceSavedAt": mtime.strftime("%Y-%m-%d %H:%M"),
    }


def _analyze_daily(wb, sections, warn, smap, sector_of, mtime):
    # ── 01 랩 ──
    wrap_tickers: set[str] = set()
    if all(s in wb.sheetnames for s in WRAP_SHEETS):
        A = _read_wrap_daily(wb[WRAP_SHEETS[0]], WRAP_SHEETS[0], warn)
        B = _read_wrap_daily(wb[WRAP_SHEETS[1]], WRAP_SHEETS[1], warn)
        wrap_tickers = set(A["rows"]) | set(B["rows"])

        contrib_a = {t: x["w"] * x["r"] for t, x in A["rows"].items()}
        contrib_b = {t: x["w"] * x["r"] for t, x in B["rows"].items()}
        rets = {t: x["r"] for t, x in {**B["rows"], **A["rows"]}.items()}
        active = {
            t: (A["rows"].get(t, {"w": 0.0})["w"] - B["rows"].get(t, {"w": 0.0})["w"]) * rets[t]
            for t in wrap_tickers
        }

        blocks = [_bars("종목 기여도", "단위 bp · 랩 수익률 기준 · note = 종목수익률",
                        _contrib_rows(contrib_a, rets))]
        if smap:
            blocks.append(_bars("분야별 기여 — 중분류",
                                "단위 bp · 위 = 당사, 아래(연한색) = BM",
                                _sector_rows(contrib_a, contrib_b, sector_of, 1), dual=True))
        blocks.append(_bars("초과수익 분해", "단위 bp · vs BM · 비중차 × 종목수익률",
                            _active_rows(active,
                                         {t: x["w"] for t, x in A["rows"].items()},
                                         {t: x["w"] for t, x in B["rows"].items()})))
        sections.append({
            "id": "wrap", "eyebrow": "01 · WRAP",
            "title": "한국투자 미국AI코어테크",
            "bm": f"BM 토러스글로벌성장랩 · 비중 합 당사 {A['weight_sum'] * 100:.0f}% "
                  f"· BM {B['weight_sum'] * 100:.0f}%",
            "scores": _scores("당사 랩", A["return"], B["return"]),
            "blocks": blocks,
        })
    else:
        warn.append("랩 시트를 찾지 못했습니다")

    # ── 02 펀드 ──
    if FUND_SHEET in wb.sheetnames:
        F = _read_fund_daily(wb[FUND_SHEET], warn)
        contrib_f = {t: x["wF"] * x["r"] for t, x in F["rows"].items()}
        contrib_b = {t: x["wB"] * x["r"] for t, x in F["rows"].items()}
        active = {t: contrib_f[t] - contrib_b[t] for t in F["rows"]}
        rets = {t: x["r"] for t, x in F["rows"].items()}
        sec = lambda t, lv: (F["rows"].get(t, {}).get("s1") or "기타")  # noqa: E731

        sections.append({
            "id": "fund", "eyebrow": "02 · FUND",
            "title": "글로벌전기차펀드",
            "bm": "BM 대비 · 시트 분야/비중 열 기준",
            "scores": _scores("당사 펀드", F["return"], F["bm"]),
            "blocks": [
                _bars("종목 기여도", "단위 bp · 펀드 수익률 기준 · note = 종목수익률",
                      _contrib_rows(contrib_f, rets)),
                _bars("국가별 기여도", "단위 bp · 티커 접미사 기준", _country_rows(contrib_f)),
                _bars("분야별 기여 — 분야1", "단위 bp · 위 = 당사, 아래(연한색) = BM",
                      _sector_rows(contrib_f, contrib_b, sec, 0), dual=True),
                _bars("초과수익 분해", "단위 bp · vs BM · 비중차 × 종목수익률",
                      _active_rows(active,
                                   {t: x["wF"] for t, x in F["rows"].items()},
                                   {t: x["wB"] for t, x in F["rows"].items()})),
            ],
        })

    est = _prev_business_day(mtime.date())
    if mtime.date() != date.today():
        warn.append(
            f"데일리 소스가 오늘 갱신되지 않았습니다(최종 저장 {mtime:%Y-%m-%d %H:%M}) "
            f"— 기준일을 {est.isoformat()} 로 추정"
        )
    return est.isoformat(), {
        "period": None,
        "dateLine": f"기준 {_ko_date(est.isoformat())} 종가 (추정)",
        "dateNote": f"미국 · 아시아 현지 종가 / 소스 저장 {mtime:%m.%d %H:%M}",
        "footnote": "**주석.** 엑셀 실시간 분석(수치 전용) — 시장 코멘트·스토리·관전 포인트는 "
                    "포함되지 않습니다. 랩·펀드 수익률은 시트 기재 종목의 Σ(비중 × 종목수익률)로, "
                    "비중 합이 100% 미만이면 미기재분(현금)을 0% 수익으로 본 근사치입니다. "
                    "기준일은 엑셀에 없어 소스 저장일의 직전 평일로 추정했습니다(휴장일 미반영).",
        "wrap_tickers": wrap_tickers,
    }


def _analyze_weekly(wb, sections, warn, smap, sector_of):
    period = {"start": None, "end": None}
    wrap_tickers: set[str] = set()
    resid_note = ""

    # ── 01 랩 ── (신규 편입 종목의 결측 기준가를 두 시트가 상호 보강 — 2패스)
    chains, shared = {}, {}
    for _pass in (1, 2):
        chains = {}
        for s in WRAP_SHEETS:
            if s not in wb.sheetnames:
                continue
            ch = _chain(wb[s], s, warn, prev_last=shared)
            if ch:
                chains[s] = ch
                shared.update(ch["last_prices"])

    if len(chains) == 2:
        A, B = chains[WRAP_SHEETS[0]], chains[WRAP_SHEETS[1]]
        bdays = {x["date"]: x for x in B["daily"]}
        contrib_a, contrib_b, active, rets_tot = {}, {}, {}, {}
        for x in A["daily"]:
            y = bdays.get(x["date"])
            for t in set(x["rets"]) | set(y["rets"] if y else {}):
                r = x["rets"].get(t, (y or {"rets": {}})["rets"].get(t))
                if r is None:
                    continue
                wf = x["weights"].get(t, 0) or 0
                wb_ = (y["weights"].get(t, 0) or 0) if y else 0
                contrib_a[t] = contrib_a.get(t, 0.0) + wf * r
                contrib_b[t] = contrib_b.get(t, 0.0) + wb_ * r
                active[t] = active.get(t, 0.0) + (wf - wb_) * r
                rets_tot[t] = (1 + rets_tot.get(t, 0.0)) * (1 + r) - 1
        wrap_tickers = set(contrib_a) | set(contrib_b)

        dates = [x["date"] for x in A["daily"]]
        period = {"start": A["start"], "end": dates[-1]}
        days = [{
            "label": f"{WEEKDAY_KO[date(*map(int, x['date'].split('-'))).weekday()]} "
                     f"{int(x['date'][5:7])}/{int(x['date'][8:10])}",
            "self": x["return"] * 100,
            "bm": (bdays[x["date"]]["return"] * 100) if x["date"] in bdays else 0.0,
            "spreadBp": round((x["return"] - (bdays[x["date"]]["return"]
                                              if x["date"] in bdays else 0.0)) * 10000),
        } for x in A["daily"]]

        spread = A["period_return"] - B["period_return"]
        resid = (spread - sum(active.values())) * 10000
        resid_note = f" 기여·초과수익은 일별 산술 합산이라 복리 연결값과 {resid:+.0f}bp 잔차가 있습니다."

        # 마지막 날 기여로 스코어카드 sub (경로 요약)
        blocks = [{
            "type": "path", "title": "일별 경로",
            "unit": "당사(진한색) vs BM(연한색) · 축 상단 = 플러스",
            "legend": "스프레드(당사−BM)는 각 일자 하단 배지",
            "days": days, "caption": None,
        }]
        blocks.append(_bars("주간 종목 기여도", "단위 bp · 일별 기여 합산 · note = 기간 수익률",
                            _contrib_rows(contrib_a, rets_tot)))
        if smap:
            blocks.append(_bars("분야별 주간 기여 — 중분류",
                                "단위 bp · 위 = 당사, 아래(연한색) = BM",
                                _sector_rows(contrib_a, contrib_b, sector_of, 1), dual=True))
        # 라벨(보유/미보유/오버·언더)은 최종일 비중으로 붙인다 — 기간 중 비중이 바뀌면
        # 한 라벨로 전 구간을 설명할 수 없어, 기준을 단위 문구에 명시한다.
        blocks.append(_bars("주간 초과수익 분해",
                            "단위 bp · vs BM · 일별 액티브 합산 · 라벨은 최종일 비중 기준",
                            _active_rows(active,
                                         {t: (A["daily"][-1]["weights"].get(t) or 0) for t in wrap_tickers},
                                         {t: ((bdays[dates[-1]]["weights"].get(t) or 0)
                                              if dates[-1] in bdays else 0) for t in wrap_tickers})))
        sections.append({
            "id": "wrap", "eyebrow": "01 · WRAP — WEEKLY",
            "title": "한국투자 미국AI코어테크",
            "bm": f"BM 토러스글로벌성장랩 · 기간별 비중 반영 · {len(days)}영업일 복리 연결",
            "scores": _scores("당사 랩 · 기간", A["period_return"], B["period_return"]),
            "blocks": blocks,
        })
    else:
        warn.append("랩 기간 블록을 읽지 못했습니다 — 시트의 '시작일' 블록을 확인해 주세요")

    # ── 02 펀드 ── (단일 비중 + 가격 블록)
    if FUND_SHEET in wb.sheetnames:
        ws = wb[FUND_SHEET]
        meta = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[1] is None:
                continue
            meta[str(r[1]).strip()] = {"s1": r[2], "wF": _num(r[5]) or 0.0, "wB": _num(r[6]) or 0.0}
        ch = _chain(ws, FUND_SHEET, warn)
        if ch:
            contrib_f, contrib_b, active, rets_tot = {}, {}, {}, {}
            fw = bw = 1.0
            days = []
            for x in ch["daily"]:
                fr = sum(meta.get(t, {"wF": 0})["wF"] * r for t, r in x["rets"].items())
                br = sum(meta.get(t, {"wB": 0})["wB"] * r for t, r in x["rets"].items())
                fw *= 1 + fr
                bw *= 1 + br
                days.append({
                    "label": f"{WEEKDAY_KO[date(*map(int, x['date'].split('-'))).weekday()]} "
                             f"{int(x['date'][5:7])}/{int(x['date'][8:10])}",
                    "self": fr * 100, "bm": br * 100, "spreadBp": round((fr - br) * 10000),
                })
                for t, r in x["rets"].items():
                    m = meta.get(t, {"wF": 0.0, "wB": 0.0})
                    contrib_f[t] = contrib_f.get(t, 0.0) + m["wF"] * r
                    contrib_b[t] = contrib_b.get(t, 0.0) + m["wB"] * r
                    active[t] = active.get(t, 0.0) + (m["wF"] - m["wB"]) * r
                    rets_tot[t] = (1 + rets_tot.get(t, 0.0)) * (1 + r) - 1
            sec = lambda t, lv: (meta.get(t, {}).get("s1") or "기타")  # noqa: E731
            sections.append({
                "id": "fund", "eyebrow": "02 · FUND — WEEKLY",
                "title": "글로벌전기차펀드",
                "bm": f"단일 비중 기준 · {len(days)}영업일 복리 연결",
                "scores": _scores("당사 펀드 · 기간", fw - 1, bw - 1),
                "blocks": [
                    {"type": "path", "title": "일별 경로",
                     "unit": "당사(진한색) vs BM(연한색)", "days": days, "caption": None},
                    _bars("주간 종목 기여도", "단위 bp · 일별 기여 합산 · note = 기간 수익률",
                          _contrib_rows(contrib_f, rets_tot)),
                    _bars("국가별 주간 기여도", "단위 bp · 티커 접미사 기준",
                          _country_rows(contrib_f)),
                    _bars("분야별 주간 기여 — 분야1", "단위 bp · 위 = 당사, 아래(연한색) = BM",
                          _sector_rows(contrib_f, contrib_b, sec, 0), dual=True),
                    _bars("주간 초과수익 분해", "단위 bp · vs BM · 일별 액티브 합산",
                          _active_rows(active,
                                       {t: m["wF"] for t, m in meta.items()},
                                       {t: m["wB"] for t, m in meta.items()})),
                ],
            })
            if period["start"] is None:
                period = {"start": ch["start"], "end": ch["daily"][-1]["date"]}

    end = period["end"] or date.today().isoformat()
    line = (f"기간 {_ko_date(period['start'])} ~ {_ko_date(end)[5:]} 종가"
            if period["start"] else "기간 —")
    return end, {
        "period": period if period["start"] else None,
        "dateLine": line,
        "dateNote": "미국 · 아시아 현지 종가 / 기간별 비중 × 일별 가격 복리 연결",
        "footnote": "**주석.** 엑셀 실시간 분석(수치 전용) — 시장 코멘트·스토리·관전 포인트는 "
                    "포함되지 않습니다. 기간 수익률은 기간별 목표비중 고정 × 일별 종목수익률을 "
                    "복리 연결한 값으로, 시트 A2 의 단일 비중 SUMPRODUCT 표시값(포트 변경 타이밍 "
                    "미반영)과 다를 수 있습니다." + resid_note +
                    " 전 종목 가격이 직전일과 동일한 행(주말·휴장·미갱신)은 자동 제외했습니다.",
        "wrap_tickers": wrap_tickers,
    }
