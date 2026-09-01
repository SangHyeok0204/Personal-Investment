"""[종목 모니터링 · 미국] 어닝 — 어닝모니터 마스터 원장 판독 (2026-09-01).

원천은 S: 의 어닝모니터 파이프라인(`raw\\모니터링\\실시간 모니터링\\어닝모니터`)이
쓰는 마스터 원장 `보유종목정리.xlsx` 의 `실적` 시트 한 장이다. 그 파이프라인의
daily-server 가 ET 08:30(BMO)·17:30(AMC) 슬롯마다 Investing.com 크롤 → SEC 8-K
구조화 → Claude CLI 관전포인트/시장반응 생성까지 돌려 이 시트에 UPSERT 한다.

**왜 여기서 크롤링하지 않는가**: 그 파이프라인은 Playwright 크롤 + SEC EDGAR +
`claude` CLI 서브프로세스를 쓰는데 이 컨테이너에는 claude 가 없고 인증도 구독
OAuth 라 옮길 수 없다. 텔레그램·매크로·성과보고와 같은 "S: 가 굽고 대시보드는
읽는다" 배선이다. Slack 알림도 계속 그쪽 서버가 쏜다 — 대시보드는 읽기만 한다.

**실시간성**: 마스터는 tempfile→`os.replace` 로 원자적으로 저장된다
(어닝모니터 `src/earnings_store.py:275-305`). 그래서 mtime 이 바뀐 판만 다시
파싱하면 반쪽 파일을 읽을 일이 없다. 평시 한 사이클은 stat 한 번(SMB 왕복 1회)이고
파싱은 파일이 실제로 바뀐 판에만 돈다. 미장은 발표가 KST 새벽·저녁에 몰려 장중엔
바뀔 일이 없지만, 앞으로 한국·중국 탭이 같은 배선을 쓸 것이라 주기는 분 아래로 잡는다.

⚠️사람이 이 xlsx 를 Excel 로 열어 저장하면 S드라이브 DRM(DOCUMENT SAFER)에 래핑돼
  openpyxl 이 못 연다. 그때는 직전에 성공한 판을 그대로 들고 note 만 세운다 —
  화면이 아무 말 없이 비는 것이 제일 나쁘다.
"""
from __future__ import annotations

import hashlib
import io
import json
import os
import sys
import threading
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo

KST = timezone(timedelta(hours=9))
ET = ZoneInfo("America/New_York")

MASTER_PATH = os.environ.get(
    "EARNINGS_MASTER_XLSX", "/srv/legacy/earnings/보유종목정리.xlsx"
)
SHEET = "실적"
META_SHEET = "메타"

# 결과 목록 창. 분기 실적이라 한 종목이 이 창에 두 번 들어올 일은 없고(다음 발표는
# 석 달 뒤), 45일이면 직전 실적 시즌(보통 4~6주)이 통째로 들어온다.
RESULT_WINDOW_DAYS = 45
# 예정 목록 창. 뒤로 하루를 여는 이유: 파이프라인은 결과를 채우면서 예정일을 비우므로,
# 어제 AMC 발표는 결과가 도착하기 전까지 '예정'에 남아 있어야 한다.
UPCOMING_BACK_DAYS = 1
UPCOMING_FORWARD_DAYS = 21

# 크롤링 서버 heartbeat = 메타 시트 `daily_server_*` 중 가장 최근 시각. 이보다 오래
# 조용하면 서버가 멎은 것으로 본다. 주말엔 발표가 없어 슬롯이 통째로 비므로 하루치
# 침묵으로 경보를 세우면 월요일마다 오탐이 난다 — 그래서 이틀이다.
HEARTBEAT_STALE_H = 48

# 관전포인트는 마스터에서 1~4 번으로 열이 갈려 있다(관전포인트N / 관전포인트결과N).
WATCHPOINT_SLOTS = (1, 2, 3, 4)

# ★★마스터는 티커당 **한 행**에 '다음 발표'와 '직전 발표'를 같이 담는다. 그래서 한 행의
#   열들이 서로 다른 분기를 가리킨다 — 이 모듈에서 제일 틀리기 쉬운 자리다.
#     · 예정실적발표일·Before/After·시가총액·예상매출·예상EPS·관전포인트1~4
#         → `upsert_forecasts` 가 **다음 발표** 기준으로 덮어쓴다(날짜가 바뀌면 관전포인트도 비운다).
#     · 결과매출·결과EPS·컨센대비·관전포인트결과1~4·기타시장반응·분기·핵심지표
#         → 아직 **직전 발표** 값이다. 새 결과가 도착해야 갈린다.
#   근거: 어닝모니터 `src/earnings_store.py:869-889`(forecasts) · `:944-964`(results).
#   그러니 예정 목록에 결과 열을 실으면 **지난 분기 숫자를 다음 분기 것처럼** 보여주게 된다.
#   `_upcoming_row` / `_result_row` 가 열을 서로 겹치지 않게 갈라 가져가는 이유다.
#   한 종목이 두 목록에 동시에 뜨는 경우(그러면 예상 열이 다음 분기 것으로 덮여 결과 행의
#   컨센서스가 어긋난다)는 실적이 분기(≈91일) 주기라 45일 창에서는 나오지 않는다
#   — 2026-09-01 실측 겹침 0건.

SESSION_MAP = {"After Market": "AMC", "Before Market": "BMO"}


def _log(msg: str) -> None:
    print(f"[collector] earnings: {msg}", file=sys.stderr, flush=True)


# ── 셀 → 값 ────────────────────────────────────────────────────────────────
def _s(v) -> str | None:
    """빈 셀·NaN·문자열 'nan'(파이프라인이 str(NaN) 을 쓴 자리)을 전부 None 으로."""
    if v is None:
        return None
    if isinstance(v, float) and v != v:  # NaN
        return None
    text = v.strip() if isinstance(v, str) else str(v).strip()
    return text if text and text.lower() != "nan" else None


def _date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    text = _s(v)
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _num(v) -> float | None:
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return None if v != v else float(v)
    text = _s(v)
    if not text:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def _money(v) -> str | None:
    """'2.71B' · '$57.63B' — 이미 사람이 읽는 꼴이라 통화기호만 떼고 그대로 쓴다."""
    text = _s(v)
    return text.lstrip("$").strip() if text else None


# ── 파싱 ───────────────────────────────────────────────────────────────────
def _row(cells: tuple, idx: dict[str, int]) -> dict | None:
    """마스터 한 행을 그대로 담는다. 분기를 가르는 일은 `_upcoming_row`/`_result_row` 몫."""

    def col(name: str):
        i = idx.get(name)
        return cells[i] if i is not None and i < len(cells) else None

    ticker = _s(col("티커"))
    if not ticker:
        return None

    funds_raw = _s(col("보유펀드")) or ""
    return {
        "ticker": ticker,
        "name": _s(col("기업명")) or ticker,
        "active": _s(col("관리상태")) == "활성",
        "funds": funds_raw.split(),
        "highlight": _s(col("highlight기업여부")) == "Y",
        "session": SESSION_MAP.get(_s(col("Before/After")) or ""),
        "marketCap": _money(col("시가총액")),
        # 다음 발표 쪽 열
        "scheduledOn": _date(col("예정실적발표일")),
        "epsEstimate": _num(col("예상EPS")),
        "revenueEstimate": _money(col("예상매출")),
        "points": [_s(col(f"관전포인트{n}")) for n in WATCHPOINT_SLOTS],
        "forecastUpdatedAt": _s(col("예상업데이트일시")),
        # 직전 발표 쪽 열
        "reportedOn": _date(col("최근실적발표일")),
        "quarter": _s(col("분기")),
        "epsActual": _num(col("결과EPS")),
        "revenueActual": _money(col("결과매출")),
        "consensus": _s(col("컨센대비")),
        "keyMetric": _s(col("핵심지표")),
        "keyMetricResult": _s(col("핵심지표결과")),
        "pointResults": [_s(col(f"관전포인트결과{n}")) for n in WATCHPOINT_SLOTS],
        "reaction": _s(col("기타시장반응")),
        "resultUpdatedAt": _s(col("결과업데이트일시")),
    }


def _common(row: dict) -> dict:
    return {
        "ticker": row["ticker"],
        "name": row["name"],
        "active": row["active"],
        "funds": row["funds"],
        "highlight": row["highlight"],
        "session": row["session"],
        "marketCap": row["marketCap"],
    }


def _upcoming_row(row: dict) -> dict:
    """예정 카드 — 결과 열은 **싣지 않는다**(지난 분기 값이라 다음 분기 것처럼 보인다)."""
    return {
        **_common(row),
        "date": _iso(row["scheduledOn"]),
        "epsEstimate": row["epsEstimate"],
        "revenueEstimate": row["revenueEstimate"],
        "watchpoints": [{"point": p} for p in row["points"] if p],
        "updatedAt": row["forecastUpdatedAt"],
    }


def _result_row(row: dict) -> dict:
    """결과 카드 — 관전포인트와 그 결과를 같은 자리에 짝지어 낸다.

    결과가 아직 안 붙은 관전포인트(LLM 분석이 하루 늦게 온다)는 `result: null` 로 남긴다.
    """
    return {
        **_common(row),
        "date": _iso(row["reportedOn"]),
        "quarter": row["quarter"],
        "epsEstimate": row["epsEstimate"],
        "epsActual": row["epsActual"],
        "revenueEstimate": row["revenueEstimate"],
        "revenueActual": row["revenueActual"],
        "consensus": row["consensus"],
        "keyMetric": row["keyMetric"],
        "keyMetricResult": row["keyMetricResult"],
        "watchpoints": [
            {"point": p, "result": r}
            for p, r in zip(row["points"], row["pointResults"])
            if p
        ],
        "reaction": row["reaction"],
        "updatedAt": row["resultUpdatedAt"],
    }


def _parse(path: str) -> dict:
    """마스터 한 판을 통째로 읽는다. 실패는 호출자가 직전 판으로 버틴다.

    바이트를 먼저 읽어 BytesIO 로 여는 것은 price_returns 와 같은 이유다 — SMB 위의
    파일 핸들을 오래 잡고 있지 않는다.
    """
    import openpyxl  # 지연 import — 순수 파싱부 테스트가 openpyxl 없이 돌도록

    with open(path, "rb") as f:
        blob = f.read()
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    try:
        if SHEET not in wb.sheetnames:
            raise ValueError(f"'{SHEET}' 시트 없음")
        stream = wb[SHEET].iter_rows(values_only=True)
        header = next(stream, None) or ()
        idx = {name: i for i, name in enumerate(header) if isinstance(name, str)}
        if "티커" not in idx or "기업명" not in idx:
            raise ValueError(f"헤더가 예상과 다름: {list(idx)[:8]}")
        rows = [r for r in (_row(c, idx) for c in stream) if r]

        # 메타 시트(key/value/updated_at) — 크롤링 서버 heartbeat 를 여기서 뽑는다.
        meta: dict[str, str] = {}
        if META_SHEET in wb.sheetnames:
            mstream = wb[META_SHEET].iter_rows(values_only=True)
            next(mstream, None)  # 헤더
            for cells in mstream:
                key = _s(cells[0] if cells else None)
                value = _s(cells[1] if len(cells) > 1 else None)
                if key and value:
                    meta[key] = value
    finally:
        wb.close()
    return {"rows": rows, "meta": meta}


def _heartbeat(meta: dict[str, str]) -> tuple[str | None, str | None]:
    """메타의 `daily_server_*` 중 가장 최근 시각과 그 키."""
    best: tuple[datetime, str, str] | None = None
    for key, value in meta.items():
        if not key.startswith("daily_server_"):
            continue
        try:
            when = datetime.fromisoformat(value)
        except ValueError:
            continue
        if best is None or when > best[0]:
            best = (when, value, key)
    return (best[1], best[2]) if best else (None, None)


def _iso(d: date | None) -> str | None:
    return d.isoformat() if d else None


class EarningsMonitor:
    """마스터를 mtime 기준으로 읽어 payload 로 들고 있는다(telegram_news 와 같은 꼴)."""

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._payload: dict | None = None
        self._etag = ""
        self._sig: tuple[int, int] | None = None  # (mtime_ns, size)
        self._parsed: dict | None = None
        self._mtime = 0.0

    def available(self) -> bool:
        return os.path.isdir(os.path.dirname(MASTER_PATH))

    def refresh(self) -> None:
        """디스크 판독(SMB). 백그라운드 루프에서만 호출한다.

        파일이 그대로여도 payload 는 다시 만든다 — 목록을 가르는 기준이 'ET 오늘'이라
        날이 바뀌는 순간 예정/결과 경계가 움직여야 한다(파일 mtime 으로는 못 잡는다).
        무거운 xlsx 파싱만 mtime 으로 건너뛴다.
        """
        now = datetime.now(KST)
        note: str | None = None
        try:
            st = os.stat(MASTER_PATH)
        except OSError:
            if self._parsed is None:
                self._store(
                    self._empty(now, f"마스터 원장을 찾지 못했습니다 — {MASTER_PATH}")
                )
                return
            note = "마스터 원장에 못 닿았습니다 — 직전 판으로 표시 중"
        else:
            sig = (st.st_mtime_ns, st.st_size)
            if sig != self._sig:
                try:
                    self._parsed = _parse(MASTER_PATH)
                except Exception as exc:  # noqa: BLE001 - 직전 판으로 버틴다
                    _log(f"마스터 판독 실패: {exc!r}")
                    if self._parsed is None:
                        self._store(
                            self._empty(
                                now, "마스터 원장을 읽지 못했습니다(DRM 래핑 의심)."
                            )
                        )
                        return
                    note = "마스터 재판독 실패 — 직전 판으로 표시 중"
                else:
                    self._sig = sig
                    self._mtime = st.st_mtime
        self._store(self._build(self._parsed, now, note))

    # -- payload --
    def _empty(self, now: datetime, note: str) -> dict:
        return {
            "generatedAt": None,
            "readAt": now.isoformat(),
            "available": False,
            "stale": True,
            "heartbeat": None,
            "heartbeatKey": None,
            "asOfET": datetime.now(ET).date().isoformat(),
            "windowDays": RESULT_WINDOW_DAYS,
            "masterPath": MASTER_PATH,
            "note": note,
            "upcoming": [],
            "results": [],
        }

    def _build(self, parsed: dict, now: datetime, note: str | None) -> dict:
        today = datetime.now(ET).date()
        up_from = today - timedelta(days=UPCOMING_BACK_DAYS)
        up_to = today + timedelta(days=UPCOMING_FORWARD_DAYS)
        res_from = today - timedelta(days=RESULT_WINDOW_DAYS)

        upcoming = [
            r
            for r in parsed["rows"]
            if r["scheduledOn"] and up_from <= r["scheduledOn"] <= up_to
        ]
        upcoming.sort(key=lambda r: (r["scheduledOn"], r["ticker"]))

        results = [
            r for r in parsed["rows"] if r["reportedOn"] and r["reportedOn"] >= res_from
        ]
        # 날짜 내림차순 + 같은 날은 티커 오름차순. 파이썬 정렬이 안정적이라 두 번 돌린다.
        results.sort(key=lambda r: r["ticker"])
        results.sort(key=lambda r: r["reportedOn"], reverse=True)

        beat, beat_key = _heartbeat(parsed["meta"])
        stale = True
        if beat:
            try:
                gap = now - datetime.fromisoformat(beat).replace(tzinfo=KST)
                stale = gap > timedelta(hours=HEARTBEAT_STALE_H)
            except ValueError:
                stale = True

        return {
            "generatedAt": datetime.fromtimestamp(self._mtime, KST).isoformat()
            if self._mtime
            else None,
            "readAt": now.isoformat(),
            "available": True,
            "stale": stale,
            "heartbeat": beat,
            "heartbeatKey": beat_key,
            "asOfET": today.isoformat(),
            "windowDays": RESULT_WINDOW_DAYS,
            "masterPath": MASTER_PATH,
            "note": note,
            "upcoming": [_upcoming_row(r) for r in upcoming],
            "results": [_result_row(r) for r in results],
        }

    def _store(self, payload: dict) -> None:
        body = dict(payload)
        body.pop("readAt", None)
        etag = (
            '"em-'
            + hashlib.sha1(
                json.dumps(body, ensure_ascii=False, sort_keys=True).encode("utf-8")
            ).hexdigest()[:16]
            + '"'
        )
        with self._lock:
            self._payload = payload
            self._etag = etag

    def serve(self) -> tuple[dict | None, str]:
        with self._lock:
            return self._payload, self._etag


_INSTANCE: EarningsMonitor | None = None


def instance() -> EarningsMonitor:
    global _INSTANCE
    if _INSTANCE is None:
        _INSTANCE = EarningsMonitor()
    return _INSTANCE
