"""WRAP 포트폴리오 실시간 수익률 — collector 어댑테이션.

구시스템 파이프라인(wrap_source_refresh 08:50 PDF 재기록 → wrap_watchlist 15s
cycle → wrap.js emit)을 읽기전용 컨테이너에 맞게 재구성한다:

  * S: 는 전부 :ro — PDF/JS/CSV 를 일절 기록하지 않는다 (read-only invariant).
  * 운용역 소스 xlsx(``이상 포트폴리오 수익률.xlsx``)를 **직접** 읽어 보유목록을
    메모리에서 구성한다. 검증 규칙(기준일 신선도·비중 수치·합·스케일 판별)은
    ``wrap_source_refresh.refresh_portfolio_pdf`` 와 동일하게 적용하고, 검증
    실패 시 구시스템이 마지막으로 기록한 포트폴리오 PDF(xlsx)로 폴백한다.
  * 분류(대/중/소)는 소스 xlsx 의 ``종목_분류`` 시트에서 직접 읽고(mtime 캐시),
    실패 시 classification.json 폴백.
  * 티커 해소·가격 수신은 verbatim ``WrapWatchlist`` 의 ``_resolve`` /
    ``_fetch_prices`` 를 재사용한다 (KisMaster + rest.snapshots).

페이로드 스키마는 구 wrap.js(window.__WRAP__)와 동일 + 부가 필드
(``holdings_source``, ``basis_date``)만 추가.
"""
from __future__ import annotations

import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from collector.wrap_returns import expected_basis_date
from etf_inav.data_sources.wrap_classification import (
    _read_classification_sheet,
    load_classification,
)
from etf_inav.data_sources.wrap_source_refresh import (
    MAX_BASIS_AGE_DAYS,
    WEIGHT_SUM_MIN,
    WEIGHT_SUM_TOLERANCE,
    _is_num,
    _read_source,
)
from etf_inav.data_sources.wrap_watchlist import (
    CASH_TICKERS,
    WrapWatchlist,
    _read_pdf_rows,
)
from kis_api.identifier import KIS_EXCHANGE_CURRENCY

WRAP_CONFIG_DIR = Path(
    os.environ.get("COLLECTOR_WRAP_CONFIG", "/srv/legacy/wrap_config")
)
WRAP_DATA_DIR = Path(os.environ.get("COLLECTOR_WRAP_DATA", "/srv/legacy/wrap_data"))
WRAP_SOURCE_DIR = Path(
    os.environ.get("COLLECTOR_WRAP_SOURCE", "/srv/legacy/wrap_source")
)

# 소스/config 재확인 주기(초). 실제 xlsx 재읽기는 mtime 변경 시에만.
SOURCE_CHECK_S = 60.0

# config.json 의 Windows 절대경로 → 컨테이너 마운트 경로 접두사 맵.
_WIN_PREFIX_MAP: list[tuple[str, Path]] = [
    ("S:\\GE\\raw\\data\\TORUS_코어테크_모니터", WRAP_DATA_DIR),
    (
        "S:\\GE\\Wonjae\\02_운용펀드_글로벌\\(Wrap) 한국투자 미국 AI코어테크랩\\4_수익률",
        WRAP_SOURCE_DIR,
    ),
]


def _log(msg: str) -> None:
    print(f"[wrap] {msg}", file=sys.stderr, flush=True)


def _parse_rebal_date(v) -> str | None:
    """``"2026.03.10 기준"`` 또는 datetime → ``"YYYY-MM-DD"`` (파싱 불가 시 None)."""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    m = re.match(r"\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", str(v))
    if not m:
        return None
    y, mo, d = (int(x) for x in m.groups())
    return f"{y:04d}-{mo:02d}-{d:02d}"


# ── 리밸 전후 기여도 계산 (Price 시계열 + 종목명→티커) ──────────────────
# Price 시트 구조는 wrap_returns 와 동일: D=날짜, 2행 E~=티커, 3행~=일별 종가.
_PRICE_DATE_COL = 4          # D (1-indexed)
_PRICE_FIRST_TICKER_COL = 5  # E
_PRICE_TICKER_ROW = 2
_PRICE_MAX_COL = 74          # BV
_PRICE_MAX_ROW = 400
_REBAL_WINDOW_TD = 5         # 리밸 전/후 '1주' = 5 거래일


def _read_price_series(wb) -> tuple[list, dict] | None:
    """Price 시트 → (dates[date 오름차순], {TICKER: [close|None, dates 정렬]})."""
    if "Price" not in wb.sheetnames:
        return None
    ws = wb["Price"]
    rows = list(
        ws.iter_rows(min_row=1, max_row=_PRICE_MAX_ROW, max_col=_PRICE_MAX_COL,
                     values_only=True)
    )
    if len(rows) < 3:
        return None
    header = rows[_PRICE_TICKER_ROW - 1]
    tick_cols: dict[int, str] = {}
    for ci in range(_PRICE_FIRST_TICKER_COL - 1, len(header)):
        t = header[ci]
        if isinstance(t, str) and t.strip():
            tick_cols[ci] = t.strip().upper()
    if not tick_cols:
        return None
    di = _PRICE_DATE_COL - 1
    dates: list = []
    series: dict[str, list] = {tk: [] for tk in tick_cols.values()}
    for rv in rows[2:]:
        if len(rv) <= di or rv[di] is None:
            continue
        d = rv[di]
        d = d.date() if hasattr(d, "date") else None
        if d is None:
            continue
        dates.append(d)
        for ci, tk in tick_cols.items():
            v = rv[ci] if ci < len(rv) else None
            series[tk].append(float(v) if isinstance(v, (int, float)) and v > 0 else None)
    return (dates, series) if dates else None


def _read_name_ticker(wb) -> dict[str, str]:
    """종목_분류 시트 → {종목명: 티커} (리밸 시트 종목명 → Price 티커 브리지)."""
    if "종목_분류" not in wb.sheetnames:
        return {}
    ws = wb["종목_분류"]
    out: dict[str, str] = {}
    for r in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        nm = r[0] if len(r) > 0 else None
        tk = r[1] if len(r) > 1 else None
        if nm and tk:
            out[str(nm).strip()] = str(tk).strip().upper()
    return out


def _asof_idx(dates: list, d) -> int:
    """dates(오름차순)에서 d 이하인 마지막 인덱스(94개라 선형)."""
    idx = -1
    for i, dd in enumerate(dates):
        if dd <= d:
            idx = i
        else:
            break
    return idx


def _window_contrib(dates, series, name2ticker, holdings, i0, i1) -> dict:
    """holdings(리밸 시점 목표비중) 기준 [i0,i1] 구간 종목별 기여도.

    ret = close[i1]/close[i0]-1, contrib = 비중 × ret. 가격 없는 종목은 제외
    (priced_n/total_n 로 커버리지 노출). top(기여 상위 종목)·cats(대분류 집계) 반환.
    """
    stocks: list[dict] = []
    cat_map: dict[str, float] = {}
    ret_total = 0.0
    priced_n = 0
    for h in holdings:
        tk = name2ticker.get(h["name"])
        s = series.get(tk) if tk else None
        if not s:
            continue
        c0 = s[i0] if 0 <= i0 < len(s) else None
        c1 = s[i1] if 0 <= i1 < len(s) else None
        if not (c0 and c1):
            continue
        ret = (c1 / c0 - 1.0) * 100.0
        w = (h["weight_pct"] or 0.0) / 100.0
        contrib = w * ret
        ret_total += contrib
        priced_n += 1
        cat = h["cat1"] or "기타"
        cat_map[cat] = cat_map.get(cat, 0.0) + contrib
        stocks.append(
            {
                "name": h["name"],
                "cat1": h["cat1"],
                "weight_pct": h["weight_pct"],
                "ret_pct": round(ret, 4),
                "contrib_pct": round(contrib, 4),
            }
        )
    stocks.sort(key=lambda x: x["contrib_pct"], reverse=True)
    cats = sorted(
        [{"cat1": k, "contrib_pct": round(v, 4)} for k, v in cat_map.items()],
        key=lambda x: x["contrib_pct"],
        reverse=True,
    )
    return {
        "ret_total": round(ret_total, 4),
        "priced_n": priced_n,
        "total_n": len(holdings),
        "top": stocks[:5],
        "cats": cats,
    }


def map_windows_path(win_path: str) -> Path | None:
    """config.json 의 ``S:\\...`` 절대경로를 컨테이너 :ro 마운트 경로로 변환."""
    raw = str(win_path or "").strip()
    if not raw:
        return None
    for prefix, root in _WIN_PREFIX_MAP:
        if raw.lower().startswith(prefix.lower()):
            rest = raw[len(prefix):].lstrip("\\/")
            return root / rest.replace("\\", "/") if rest else root
    return None


def holdings_from_source(
    src_path: Path, sheet: str
) -> tuple[list[dict] | None, str, dict]:
    """운용역 소스 시트 → wrap_watchlist 보유행(list) 를 메모리에서 구성.

    검증 규칙은 ``wrap_source_refresh.refresh_portfolio_pdf`` 와 동일 (파일
    기록·sidecar 만 없음). 실패 시 (None, 사유, {}) — 호출부가 PDF 폴백.
    세 번째 원소는 ① 종가수익률에 쓰는 T-2 기준일·환율(extra).
    """
    rows, basis_date, extra = _read_source(src_path, sheet)
    if basis_date is None:
        return None, "no_basis_date(F2)", {}
    if (datetime.now().date() - basis_date).days > MAX_BASIS_AGE_DAYS:
        return None, f"stale_basis({basis_date})", {}

    securities = [r for r in rows if not r["is_cash"]]
    cash_rows = [r for r in rows if r["is_cash"]]
    if not securities:
        return None, "no_securities", {}
    bad = [r["ticker"] or r["name"] for r in securities if not _is_num(r["weight"])]
    if bad:
        return None, f"non_numeric_weight(n={len(bad)})", {}
    cash_w_raw = sum(r["weight"] for r in cash_rows if _is_num(r["weight"]))
    sec_w_raw = sum(r["weight"] for r in securities)
    total_raw = sec_w_raw + cash_w_raw

    # 스케일 판별: 분수(합 0.75~1.05) → ×100, 퍼센트(합 75~105) → ×1, 그 외 거부.
    if WEIGHT_SUM_MIN / 100.0 <= total_raw <= 1.0 + 0.05:
        factor = 100.0
    elif WEIGHT_SUM_MIN <= total_raw <= 100.0 + 5.0:
        factor = 1.0
    else:
        return None, f"weight_sum_off({total_raw:.4f})", {}
    final_total = total_raw * factor
    if not (WEIGHT_SUM_MIN <= final_total <= 100.0 + WEIGHT_SUM_TOLERANCE):
        return None, f"weight_sum_off({final_total:.3f}%)", {}

    out: list[dict] = []
    for r in securities:
        prev = r["prev"] if (_is_num(r["prev"]) and r["prev"] > 0) else None
        prev2 = r["prev2"] if (_is_num(r["prev2"]) and r["prev2"] > 0) else None
        out.append(
            {
                "ticker": r["ticker"],
                "name": "" if r["name"] is None else str(r["name"]),
                "weight_pct": round(r["weight"] * factor, 6),
                "prev2_close": prev2,
                "prev_close": prev,
                "exchange": None,  # WrapWatchlist 가 KisMaster 로 자동 해소
            }
        )
    if cash_rows and _is_num(cash_rows[0]["weight"]):
        out.append(
            {
                "ticker": "CASH",
                "name": "현금",
                "weight_pct": round(cash_w_raw * factor, 6),
                "prev2_close": None,
                "prev_close": None,
                "exchange": None,
            }
        )
    return out, basis_date.strftime("%Y%m%d"), extra


class WrapCollector:
    def __init__(self, master, rest, us_daytime_mode: str = "auto") -> None:
        self.watchlist = WrapWatchlist(master, rest, us_daytime_mode)
        self.config: dict = {}
        self._config_mtime: float | None = None
        self._holdings: dict[str, list[dict]] = {}
        self._holdings_meta: dict[str, dict] = {}
        self._last_source_check = 0.0
        self._cls_cache: dict = {}
        self._cls_mtime: float | None = None
        # ── 성과 비교(track record): Port_Bloommberg/Port_TORUS 누적수익률% 시계열 ──
        self._perf_cache: dict | None = None
        self._perf_mtime: float | None = None
        # ── 리밸런싱 이력(track record): 리밸런싱_히스토리 시점별 편입 구성 ──
        self._rebal_cache: dict | None = None
        self._rebal_mtime: float | None = None

    # ── 성과 비교 시계열 (track record: 자사 vs TORUS 누적수익률%) ──────
    def build_performance(self) -> dict | None:
        """Port_Bloommberg(자사 AI코어테크랩)·Port_TORUS(BM) 시트의 A=날짜·B=누적수익률%
        시계열을 읽어 반환. mtime 캐시. 컨테이너는 이 xlsx 를 평문(PK)으로 읽는다
        (로컬 Windows 의 DOCUMENT SAFER 표시와 무관)."""
        import openpyxl

        cfg = self._load_config()
        if not cfg:
            return None
        # 소스 xlsx 위치 = wrap_sources 중 아무 경로나 (동일 '이상 …' 파일).
        src_path = None
        for s in (cfg.get("wrap_sources") or {}).values():
            p = map_windows_path((s or {}).get("path", ""))
            if p is not None and p.exists():
                src_path = p
                break
        if src_path is None:
            return self._perf_cache
        try:
            mtime = src_path.stat().st_mtime
        except OSError:
            mtime = None
        if self._perf_cache is not None and self._perf_mtime == mtime:
            return self._perf_cache

        sheets = {
            "aicoretech": ("Port_Bloommberg", "AI코어테크랩"),
            "torus": ("Port_TORUS", "TORUS"),
        }
        series: dict[str, dict] = {}
        try:
            wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
            try:
                for key, (sheet, label) in sheets.items():
                    if sheet not in wb.sheetnames:
                        continue
                    ws = wb[sheet]
                    pts: list[list] = []
                    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2,
                                            values_only=True):
                        d, ret = row[0], row[1]
                        if d is None:
                            break
                        ds = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10]
                        if not isinstance(ret, (int, float)):
                            continue
                        pts.append([ds, round(float(ret), 6)])
                    if pts:
                        series[key] = {
                            "label": label,
                            "base_date": pts[0][0],
                            "last_date": pts[-1][0],
                            "points": pts,
                        }
            finally:
                wb.close()
        except Exception as exc:  # noqa: BLE001 - last-good 유지
            _log(f"performance 읽기 실패: {exc!r}")
            return self._perf_cache

        payload = {
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "series": series,
        }
        self._perf_cache = payload
        self._perf_mtime = mtime
        return payload

    # ── 리밸런싱 이력 (track record: 자사·TORUS 시점별 편입 구성) ────────
    def build_rebalancing(self) -> dict | None:
        """``리밸런싱_히스토리`` 시트 → 포트폴리오별 리밸 시점 편입 구성 이력.

        와이드 포맷: ``항목`` 헤더행마다 한 포트폴리오 블록(등장순 = AI코어테크랩,
        TORUS). 각 리밸 시점은 5열 ``[종목·대분류·중분류·소분류·비중]`` 블록으로
        열방향 나열되고, 헤더행 각 블록 첫 열에 ``"YYYY.MM.DD 기준"`` 날짜가 있다.
        종목행은 비중 내림차순(블록마다 개수 상이 → 하단은 공백)이며 ``기타/(미분류)/
        (현금)`` 은 블록 하단 고정행. 비중은 분수(0.1) → % 로 변환. build_performance
        와 동일한 mtime 캐시·평문(PK) 판독(로컬 DOCUMENT SAFER 표시와 무관)."""
        import openpyxl

        cfg = self._load_config()
        if not cfg:
            return None
        src_path = None
        for s in (cfg.get("wrap_sources") or {}).values():
            p = map_windows_path((s or {}).get("path", ""))
            if p is not None and p.exists():
                src_path = p
                break
        if src_path is None:
            return self._rebal_cache
        try:
            mtime = src_path.stat().st_mtime
        except OSError:
            mtime = None
        if self._rebal_cache is not None and self._rebal_mtime == mtime:
            return self._rebal_cache

        sheet_name = "리밸런싱_히스토리"
        try:
            wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 - last-good 유지
            _log(f"rebalancing 로드 실패: {exc!r}")
            return self._rebal_cache
        try:
            if sheet_name not in wb.sheetnames:
                return self._rebal_cache
            ws = wb[sheet_name]
            grid = [
                list(r)
                for r in ws.iter_rows(
                    min_row=1, max_row=ws.max_row,
                    min_col=1, max_col=ws.max_column, values_only=True,
                )
            ]
            # 리밸 전/후 기여도 계산용 (같은 워크북에서 함께 판독).
            price = _read_price_series(wb)
            name2ticker = _read_name_ticker(wb)
        finally:
            wb.close()

        def cell(r: int, c: int):  # 1-indexed, 범위 밖 None
            if 1 <= r <= len(grid) and 1 <= c <= len(grid[r - 1]):
                return grid[r - 1][c - 1]
            return None

        # 포트폴리오 블록 경계 = col A 가 ``항목`` 인 헤더행. 등장순으로 매핑.
        header_rows = [
            r for r in range(1, len(grid) + 1)
            if str(cell(r, 1) or "").strip() == "항목"
        ]
        order = [("aicoretech", "AI코어테크랩"), ("torus", "TORUS")]
        portfolios: dict[str, dict] = {}
        for idx, hr in enumerate(header_rows):
            if idx >= len(order):
                break
            key, label = order[idx]
            data_start = hr + 2  # 헤더행 + 서브헤더행 다음
            data_end = (
                header_rows[idx + 1] - 1 if idx + 1 < len(header_rows) else len(grid)
            )
            # 블록은 세로로 두 섹션: ①"편입 종목 및 비중"(개별 종목+현금),
            # ②"분류별 비중"(대분류 소계). col A 의 "분류별" 라벨이 ② 시작 경계.
            sec2_hr = None
            for r in range(hr + 1, data_end + 1):
                if "분류별" in str(cell(r, 1) or ""):
                    sec2_hr = r
                    break
            sec1_end = (sec2_hr - 1) if sec2_hr else data_end

            def _wp(r: int, c: int) -> float:
                w = cell(r, c + 4)
                return round(float(w) * 100.0, 4) if isinstance(w, (int, float)) else 0.0

            events: list[dict] = []
            c = 2  # 첫 블록 시작 열 (stride 5)
            while True:
                ds = _parse_rebal_date(cell(hr, c))
                if ds is None:
                    break
                # ① 개별 종목 + 현금 (종목수는 시점마다 달라 하단 공백 → skip)
                holdings: list[dict] = []
                cash = 0.0
                for r in range(data_start, sec1_end + 1):
                    nm = cell(r, c)
                    if nm is None or str(nm).strip() == "":
                        continue
                    name = str(nm).strip()
                    if name == "(현금)":
                        cash = _wp(r, c)
                        continue
                    holdings.append(
                        {
                            "name": name,
                            "cat1": str(cell(r, c + 1) or "").strip(),
                            "cat2": str(cell(r, c + 2) or "").strip(),
                            "cat3": str(cell(r, c + 3) or "").strip(),
                            "weight_pct": _wp(r, c),
                        }
                    )
                # ② 대분류 소계 + 기타/미분류
                cats: list[dict] = []
                etc = unc = 0.0
                if sec2_hr is not None:
                    for r in range(sec2_hr + 1, data_end + 1):
                        nm = cell(r, c)
                        if nm is None or str(nm).strip() == "":
                            continue
                        name = str(nm).strip()
                        if name == "(현금)":
                            continue
                        if name == "기타":
                            etc = _wp(r, c)
                            continue
                        if name == "(미분류)":
                            unc = _wp(r, c)
                            continue
                        cats.append({"name": name, "weight_pct": _wp(r, c)})
                events.append(
                    {
                        "date": ds,
                        "holdings": holdings,
                        "n_holdings": len(holdings),
                        "cash_pct": cash,
                        "etc_pct": etc,
                        "unclassified_pct": unc,
                        "cats": cats,
                    }
                )
                c += 5
            events.sort(key=lambda e: e["date"])
            # 리밸 전/후 1주 기여도 enrich (Price 시계열 있을 때만).
            if price is not None:
                dates, series = price
                for i, ev in enumerate(events):
                    try:
                        d = datetime.strptime(ev["date"], "%Y-%m-%d").date()
                    except ValueError:
                        ev["perf"] = {"before": None, "after": None}
                        continue
                    idx = _asof_idx(dates, d)
                    after = before = None
                    if idx >= 0:
                        i1 = min(idx + _REBAL_WINDOW_TD, len(dates) - 1)
                        if i1 > idx:  # 리밸 후 1주: 신규 구성 비중
                            after = _window_contrib(
                                dates, series, name2ticker, ev["holdings"], idx, i1
                            )
                            after["start"] = dates[idx].strftime("%Y-%m-%d")
                            after["end"] = dates[i1].strftime("%Y-%m-%d")
                        i0 = max(idx - _REBAL_WINDOW_TD, 0)
                        if i > 0 and i0 < idx:  # 리밸 전 1주: 직전 구성 비중
                            before = _window_contrib(
                                dates, series, name2ticker,
                                events[i - 1]["holdings"], i0, idx,
                            )
                            before["start"] = dates[i0].strftime("%Y-%m-%d")
                            before["end"] = dates[idx].strftime("%Y-%m-%d")
                    ev["perf"] = {"before": before, "after": after}
            portfolios[key] = {"label": label, "events": events}

        payload = {
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "portfolios": portfolios,
        }
        self._rebal_cache = payload
        self._rebal_mtime = mtime
        return payload

    # ── config / 보유목록 / 분류 ───────────────────────────────────────
    def _load_config(self) -> dict:
        path = WRAP_CONFIG_DIR / "config.json"
        try:
            mtime = path.stat().st_mtime
        except OSError:
            return self.config
        if self.config and self._config_mtime == mtime:
            return self.config
        try:
            self.config = json.loads(path.read_text(encoding="utf-8"))
            self._config_mtime = mtime
        except Exception as exc:  # noqa: BLE001 - last-good 유지
            _log(f"config 읽기 실패: {exc!r}")
        return self.config

    def _refresh_holdings(self, cfg: dict) -> None:
        now = time.time()
        if self._holdings and now - self._last_source_check < SOURCE_CHECK_S:
            return
        self._last_source_check = now
        pdfs = cfg.get("portfolio_pdfs", {})
        sources = cfg.get("wrap_sources", {})
        keys = list(cfg.get("portfolios") or []) or list(pdfs.keys())
        for key in keys:
            src = sources.get(key) or {}
            src_path = map_windows_path(src.get("path", ""))
            sheet = src.get("sheet", "")
            loaded: list[dict] | None = None
            meta: dict = {}
            if src_path is not None and sheet and src_path.exists():
                try:
                    src_mtime = src_path.stat().st_mtime
                except OSError:
                    src_mtime = None
                prev = self._holdings_meta.get(key) or {}
                if (
                    prev.get("source") == "SOURCE"
                    and prev.get("src_mtime") == src_mtime
                    and key in self._holdings
                ):
                    continue  # 소스 무변경 — 기존 보유목록 유지
                try:
                    holdings, basis, extra = holdings_from_source(src_path, sheet)
                except Exception as exc:  # noqa: BLE001
                    holdings, basis, extra = None, f"read_error: {exc}", {}
                if holdings:
                    loaded = holdings
                    meta = {
                        "source": "SOURCE",
                        "basis": basis,
                        "src_mtime": src_mtime,
                        "extra": extra,
                    }
                    _log(f"{key}: 소스 적용 basis={basis} n={len(holdings)}")
                else:
                    _log(f"{key}: 소스 검증 실패({basis}) — PDF 폴백")
            if loaded is None and key not in self._holdings:
                pdf_path = map_windows_path(pdfs.get(key, ""))
                if pdf_path is not None and pdf_path.exists():
                    try:
                        loaded = _read_pdf_rows(pdf_path)
                        # 폴백 PDF 에는 T-2 종가·환율이 없다 → ① 은 이 랩만 미표시.
                        meta = {
                            "source": "PDF_FALLBACK", "basis": None,
                            "src_mtime": None, "extra": {},
                        }
                        _log(f"{key}: PDF 폴백 적용 n={len(loaded)}")
                    except Exception as exc:  # noqa: BLE001
                        _log(f"{key}: PDF 읽기 실패: {exc!r}")
            if loaded:
                self._holdings[key] = loaded
                self._holdings_meta[key] = meta

    def _classification(self, cfg: dict) -> dict:
        cs = cfg.get("classification_source") or {}
        src_path = map_windows_path(cs.get("path", ""))
        sheet = cs.get("sheet") or "종목_분류"
        if src_path is not None and src_path.exists():
            try:
                mtime = src_path.stat().st_mtime
                if self._cls_mtime == mtime and self._cls_cache:
                    return self._cls_cache
                mapping = _read_classification_sheet(src_path, sheet)
                if mapping:
                    self._cls_cache = mapping
                    self._cls_mtime = mtime
                    return mapping
            except Exception as exc:  # noqa: BLE001
                _log(f"분류 시트 읽기 실패: {exc!r}")
        cj = map_windows_path(cfg.get("classification_json", ""))
        if cj is not None:
            data = load_classification(cj)
            if data:
                return data
        return self._cls_cache

    # ── 페이로드 ───────────────────────────────────────────────────────
    def build_payload(self, fx_table: dict | None = None) -> dict | None:
        cfg = self._load_config()
        if not cfg:
            return None
        self._refresh_holdings(cfg)
        if not self._holdings:
            return None
        names = cfg.get("portfolio_names", {})
        order = cfg.get("portfolios", list(self._holdings.keys()))
        classification = self._classification(cfg)
        rows_by_pf = self._holdings
        price_by_symbol = self.watchlist._fetch_prices(rows_by_pf)

        fx_live = (fx_table or {}).get("rates", {}).get("USD")  # 실시간 USD/KRW
        expected_basis = expected_basis_date(datetime.now().date()).strftime("%Y%m%d")

        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ts = int(time.time() * 1000)
        portfolios: list[dict] = []
        for key in [k for k in order if k in rows_by_pf] + [
            k for k in rows_by_pf if k not in order
        ]:
            rows = rows_by_pf[key]
            meta = self._holdings_meta.get(key) or {}
            extra = meta.get("extra") or {}
            fx_t2, fx_t1 = extra.get("fx_t2"), extra.get("fx_t1")
            # ① 환등락 = T-2 → T-1 (소스 시트 USDKRW 행 = 블룸버그 KRW L160).
            # ② 환등락 = T-1 → 실시간 (네이버 USD/KRW).
            fx_chg1 = (fx_t1 / fx_t2 - 1.0) if (fx_t1 and fx_t2) else None
            fx_chg2 = (fx_live / fx_t1 - 1.0) if (fx_live and fx_t1) else None

            holdings: list[dict] = []
            ret1_sum = ret2_sum = mw = tw = w1 = 0.0
            nm = n1 = 0
            for h in rows:
                is_cash = str(h["ticker"]).upper() in CASH_TICKERS
                r = self.watchlist._resolved.get(h["ticker"])
                info = price_by_symbol.get(r[1]) if r else None
                live = info["last"] if info else None
                exchange = h["exchange"] or (info["exchange"] if info else None)
                # 기준통화: 현금은 달러현금으로 본다(소스 시트가 현금 행 수익률을 환등락률
                # 그대로 잡는다 — 2026-07-31 사용자 확정). 그 외는 거래소→통화 맵.
                if is_cash:
                    currency = "KRW" if str(h["ticker"]).upper() == "KRW" else "USD"
                else:
                    currency = KIS_EXCHANGE_CURRENCY.get(
                        str(exchange or "").upper(), "KRW"
                    )
                w = h["weight_pct"] or 0.0
                prev2 = h.get("prev2_close")   # E열 T-2 종가
                prev = h["prev_close"]         # F열 T-1 종가

                # 수익률      = 전전일종가 → 전일종가 (현지통화). 카드 ① 의 재료.
                # 실시간수익률 = 전일종가 → 현재가 (장중 등락). 카드 ② 의 재료.
                if is_cash:
                    ret1, rt, live, matched = 0.0, 0.0, prev, True
                else:
                    ret1 = ((prev / prev2) - 1.0) * 100.0 if (prev and prev2) else None
                    rt = (
                        ((live / prev) - 1.0) * 100.0
                        if (live is not None and prev)
                        else None
                    )
                    matched = live is not None
                # 환 반영 = (1+수익률)(1+환등락) − 1. 원금 전체가 환에 노출된다는 전제로,
                # 소스 시트의 종목별 원화수익률 관례(G·H열)와 같은 식이다.
                # 환은 전 행에 적용한다 — 이 랩들은 소스 시트가 현금까지 USD 로 잡는
                # 미국물 전용이고, currency 는 KIS 가 해소한 거래소에서 오므로 가격이
                # 안 잡힌 행만 환이 빠지면 테이블 합이 카드와 어긋난다(① 은 현재가와
                # 무관해야 한다). 명시적 원화현금(ticker=KRW)만 예외.
                fxc = 0.0 if (is_cash and currency == "KRW") else fx_chg1
                ret1_krw = (
                    ((1.0 + ret1 / 100.0) * (1.0 + fxc) - 1.0) * 100.0
                    if (ret1 is not None and fxc is not None)
                    else None
                )
                # ② 구간(전일종가→현재가)의 환 반영 짝. ① 과 같은 식·같은 현금 예외를
                # 쓰고 환등락만 fx_chg2(T-1→실시간)로 바꾼다.
                # ★분류 트리의 '환율 ON' 이 이 값을 쓴다(2026-08-27) — 웹에서 환을 다시
                #   곱하면 식이 두 곳으로 갈라지므로 서버가 계산해 내려보낸다.
                #   Σ(비중 × 이 값) = return2_krw 임을 실측 확인(Δ≈5e-9).
                fxc2 = 0.0 if (is_cash and currency == "KRW") else fx_chg2
                rt_krw = (
                    ((1.0 + rt / 100.0) * (1.0 + fxc2) - 1.0) * 100.0
                    if (rt is not None and fxc2 is not None)
                    else None
                )
                contrib = (w / 100.0 * ret1) if ret1 is not None else None

                tw += w
                if contrib is not None:
                    ret1_sum += contrib
                    w1 += w
                    if not is_cash:
                        n1 += 1
                if rt is not None:
                    ret2_sum += w / 100.0 * rt
                    mw += w
                    nm += 1
                cls = classification.get(str(h["ticker"]).upper(), {})
                holdings.append(
                    {
                        "ticker": h["ticker"],
                        "name": h["name"],
                        "exchange": exchange,
                        "currency": currency,
                        "is_cash": is_cash,
                        "weight_pct": w,
                        "prev2_close": prev2,
                        "prev_close": prev,
                        "livePrice": live,
                        # 종가-대-종가(전전일→전일). 카드 ① = Σ(이 값 × 비중).
                        "return_pct": ret1,
                        # 같은 구간에 환까지 반영 = (1+수익률)(1+환등락) − 1.
                        "return_krw_pct": ret1_krw,
                        "contribution_pct": contrib,
                        # 전일종가 → 현재가.
                        "realtime_return_pct": rt,
                        # 같은 구간에 환까지 반영 (분류 트리 '환율 ON' × 실시간).
                        "realtime_return_krw_pct": rt_krw,
                        "matched": matched,
                        "tradeTime": info["tradeTime"] if info else None,
                        "cat1": cls.get("cat1") or "",
                        "cat2": cls.get("cat2") or "",
                        "cat3": cls.get("cat3") or "",
                    }
                )
            holdings.sort(key=lambda x: x["weight_pct"] or 0, reverse=True)

            def _with_fx(ret: float, chg: float | None) -> float | None:
                """포트폴리오 수익률에 환 반영 — (1+R)(1+환등락) − 1."""
                if chg is None:
                    return None
                return ((1.0 + ret / 100.0) * (1.0 + chg) - 1.0) * 100.0

            basis = meta.get("basis")
            has1 = n1 > 0  # 종가 두 개가 다 있는 종목이 하나도 없으면 ① 미표시
            portfolios.append(
                {
                    "key": key,
                    "name": names.get(key, key),
                    # 분류 트리 루트 = ① 주식분(테이블 기여도 합과 같은 값).
                    "return_pct": ret1_sum if has1 else 0.0,
                    "matched_weight_pct": mw,
                    "total_weight_pct": tw,
                    "n_matched": nm,
                    "n_total": len(rows),
                    "holdings": holdings,
                    "holdings_source": meta.get("source"),
                    "basis_date": basis,
                    # 환등락률 원본 — 화면 툴팁 검산용.
                    "fx_currency": "USD",
                    "fx_return_pct": fx_chg1 * 100.0 if fx_chg1 is not None else None,
                    "fx_realtime_pct": fx_chg2 * 100.0 if fx_chg2 is not None else None,
                    "fx_t2": fx_t2,
                    "fx_t1": fx_t1,
                    # ② 전일종가→최근체결가 (현지통화 / 환 반영)
                    "return2_usd": ret2_sum,
                    "return2_krw": _with_fx(ret2_sum, fx_chg2),
                    # ① 전전일→전일 종가수익률 (현지통화 / 환 반영) + 기준일·신선도
                    "return1_usd": ret1_sum if has1 else None,
                    "return1_krw": _with_fx(ret1_sum, fx_chg1) if has1 else None,
                    "return1_basis_date": basis,
                    "return1_prev_date": (
                        extra["prev2_date"].strftime("%Y%m%d")
                        if extra.get("prev2_date")
                        else None
                    ),
                    "return1_is_current": bool(basis and basis == expected_basis),
                    # ① 계산에 실제 들어간 비중합 (100% 미만이면 커버 누락).
                    "ret1_weight_pct": w1 if has1 else None,
                }
            )
        if not portfolios:
            return None
        return {
            "date": datetime.now().strftime("%Y%m%d"),
            "generatedAt": generated_at,
            "timestamp": ts,
            "priceGeneratedAt": generated_at,
            "portfolios": portfolios,
            "fx": {
                "rates": (fx_table or {}).get("rates") or {},
                "fetched_at": (fx_table or {}).get("fetched_at"),
            },
        }
