"""[컴퓨팅 지수 모니터링] — GPU 렌탈 지수 판독 (2026-08-26).

[AI Key Data] 좌상단 3칸 카드의 데이터원. AI Key Data가 받아 두는 블룸버그
내보내기에서 Silicon Data 렌탈 지수를 읽어 세로 분할 차트 payload 로 접는다
(패널 = ``INDICES`` 순서 — H100 · B200 · A100).

무엇을 재는 지수인가: SDH100RT·SDB200RT 는 GPU 시간당 렌탈 단가($/GPU-hr)다.
10/5 CME 에 상장 예정인 현금정산 GPU 컴퓨트 선물의 기초지수이고(현재 상장된
GPU 선물은 0개), 밸류체인에서 이 지수가 직접 재는 층은 GPUaaS 의 매출 단가다.

⚠️단가를 한 차트에 여러 y축으로 겹치면 안 된다 — 세대별로 스케일이 2배씩 차이 나
  없는 상관을 만든다. 그래서 지수 하나에 패널 하나(자기 y축)를 준다.

★2026-08-27: 세 번째 패널이 배수(SDB200/SDH100)였는데 사용자 지시로 **A100 지수**로
  교체했다. 배수는 "어느 세대가 더 빨리 오르는가"를 봤지만, 세 세대 단가를 나란히
  놓으면 같은 걸 눈으로 읽을 수 있다. 배수가 다시 필요하면 이 파일에서 되살린다.

원천은 AI Key Data 프로젝트의 `input/raw` (`/srv/legacy/gpu_compute` :ro).
"""
from __future__ import annotations

import io
import os
from datetime import date, datetime, timedelta, timezone

_KST = timezone(timedelta(hours=9))

SRC_PATH = os.environ.get(
    "COMPUTE_INDEX_XLSX", "/srv/legacy/gpu_compute/GPU임대지수_주가_통합.xlsx"
)
# 시트 이름은 내보내기마다 다르다(구 파일 `Worksheet`/`FNC`, 통합본 `fnc`/`val`).
# 이름으로 찍지 말고 **1행에 `Security` 가 있는 첫 시트**를 쓴다 — 통합본에는 종목
# OHLCV 시트(crwv·nbis·nhn·iren)가 같이 들어 있어 worksheets[0] 가정도 위험하다.
SHEET = os.environ.get("COMPUTE_INDEX_SHEET", "")

# ── 관심 지수 (사람이 고치는 곳) ──────────────────────────────────────────────
# ticker 는 xlsx 1행의 `Security` 오른쪽 칸 문자열과 정확히 일치해야 한다.
# 블록 위치(A~C / H~J)는 하드코딩하지 않는다 — `Security` 라벨을 훑어 찾는다.
INDICES: list[dict] = [
    {"key": "h100", "ticker": "SDH100RT Index", "name": "SDH100RT", "chip": "H100"},
    {"key": "b200", "ticker": "SDB200RT Index", "name": "SDB200RT", "chip": "B200"},
    # ★2026-08-27 사용자 지시로 배수(SDB200/SDH100) 패널을 A100 지수로 교체.
    #   기초 파일에 `SDA100RT Index` 블록이 들어오면 자동으로 세 번째 패널이 된다
    #   (블록 탐색이 `Security` 라벨 기준이라 열 위치는 상관없다).
    {"key": "a100", "ticker": "SDA100RT Index", "name": "SDA100RT", "chip": "A100"},
    # ★2026-08-31 사용자 지시로 추가. 앞의 셋과 **성격이 다르다** — GPU 시간당 렌탈 단가가
    #   아니라 LLM 토큰 지수다(워크북 AE 블록, 2025-12-01~, 값이 1.0 언저리).
    #   그래서 `unit` 을 따로 준다. 같은 `$/GPU-hr` 로 적으면 1.05 를 "시간당 1달러"로
    #   읽게 되는데 그건 사실이 아니다 — 워크북이 주는 건 Currency=USD 뿐이고 분모가
    #   무엇인지(토큰 100만개? 지수 기준값?)는 원천에 적혀 있지 않다. 지어내지 않는다.
    {"key": "llmtk", "ticker": "SDLLMTK Index", "name": "SDLLMTK", "chip": "LLM 토큰",
     "unit": "index (USD)"},
]
UNIT = "$/GPU-hr"


# ── 순수 계산부 (파일 IO 없음 — 테스트는 여기를 겨눈다) ──────────────────────

def _stats(points: list[tuple[date, float]]) -> dict:
    """시작·최근·최저·최고 + 구간 변화율. points 는 날짜 오름차순."""
    vals = [v for _, v in points]
    lo_i = min(range(len(vals)), key=vals.__getitem__)
    hi_i = max(range(len(vals)), key=vals.__getitem__)
    first, last = vals[0], vals[-1]
    return {
        "start": first,
        "start_date": points[0][0].isoformat(),
        "last": last,
        "last_date": points[-1][0].isoformat(),
        "min": vals[lo_i], "min_date": points[lo_i][0].isoformat(),
        "max": vals[hi_i], "max_date": points[hi_i][0].isoformat(),
        # 전일 대비 — 지수가 일간이라 마지막 두 점이 곧 DtD.
        "chg_1d_pct": ((last / vals[-2] - 1) * 100) if len(vals) > 1 and vals[-2] else None,
        "chg_pct": ((last / first - 1) * 100) if first else None,
        "n": len(vals),
    }


def build_payload(blocks: dict[str, list[tuple[date, float]]]) -> dict:
    """티커별 시계열 → 카드 payload (INDICES 순서대로 한 패널씩).

    기초 파일에 없는 지수는 조용히 빠진다 — 블록이 늘면 패널도 는다.
    """
    out: dict = {
        "generated_at": datetime.now(_KST).strftime("%Y-%m-%d %H:%M:%S"),
        "asof": None,
        "unit": UNIT,
        "note": None,
        "series": [],
    }

    for spec in INDICES:
        pts = blocks.get(spec["ticker"]) or []
        if not pts:
            continue
        pts = sorted(pts)
        out["series"].append({
            "key": spec["key"],
            "name": spec["name"],
            "label": spec["chip"],
            # 지수마다 단위가 다를 수 있다(GPU 렌탈 = $/GPU-hr, LLM 토큰 = 지수).
            "unit": spec.get("unit", UNIT),
            "kind": "price",
            "points": [[d.isoformat(), v] for d, v in pts],
            "stats": _stats(pts),
        })

    out["asof"] = max((s["stats"]["last_date"] for s in out["series"]), default=None)
    return out


# ── xlsx 판독 (mtime+size 캐시) ──────────────────────────────────────────────
# price_returns 와 같은 idiom: 바이트를 통째로 읽어 BytesIO 로 열고, 열기 실패면
# 직전 캐시를 그대로 낸다(S: 쪽 제자리 저장과 겹치는 순간 대비).
_CACHE: dict = {"sig": None, "blocks": None}


def _pick_sheet(wb):
    """1행에 `Security` 라벨이 있는 첫 시트(= 블룸버그 지수 블록이 있는 시트)."""
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            if any(isinstance(c, str) and c.strip() == "Security" for c in row):
                return ws
            break
    return wb.worksheets[0]


def _read_blocks(path: str = SRC_PATH) -> dict[str, list[tuple[date, float]]]:
    """블룸버그 내보내기 → {ticker: [(date, px_last), ...]}.

    시트 모양: 1행 `Security | <티커>`, 7행 `Date | PX_LAST | CHG_PCT_1D`,
    8행부터 데이터(최신이 위). 지수마다 이 3열 블록이 가로로 이어 붙는다.
    ★열 인덱스를 박지 않고 `Security` 라벨 위치로 블록 시작을 찾는다 —
      지수가 추가되거나 순서가 바뀌어도 안 깨진다.
    """
    st = os.stat(path)
    sig = (st.st_mtime_ns, st.st_size)
    if _CACHE["sig"] == sig and _CACHE["blocks"] is not None:
        return _CACHE["blocks"]

    with open(path, "rb") as f:
        blob = f.read()
    try:
        import openpyxl  # 지연 import — 순수 계산부 테스트가 openpyxl 없이 돌도록

        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
        ws = wb[SHEET] if SHEET and SHEET in wb.sheetnames else _pick_sheet(wb)
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        if _CACHE["blocks"] is not None:
            return _CACHE["blocks"]
        raise

    if not rows:
        return {}

    # 1행에서 (블록 시작열, 티커) 수집 — Date 열 = 'Security' 가 있던 열.
    head = rows[0]
    starts = [(i, str(head[i + 1]).strip())
              for i, c in enumerate(head[:-1])
              if isinstance(c, str) and c.strip() == "Security" and head[i + 1]]

    blocks: dict[str, list[tuple[date, float]]] = {}
    for col, ticker in starts:
        pts: list[tuple[date, float]] = []
        for r in rows[7:]:
            if col + 1 >= len(r):
                continue
            d, v = r[col], r[col + 1]
            if not isinstance(d, datetime) or not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            pts.append((d.date(), float(v)))
        if pts:
            blocks[ticker] = sorted(pts)

    _CACHE["sig"] = sig
    _CACHE["blocks"] = blocks
    return blocks


def build_compute_index() -> dict:
    """xlsx → 카드 payload 한 장.

    ★원천 파일이 없으면 **503 이 아니라 빈 payload + note** 로 돌려준다
      (2026-08-27). 이 파일은 파이프라인 산출물이 아니라 사람이 받아 두는 블룸버그
      내보내기라 폴더에서 사라질 수 있는데, 그때 503 이 나면 카드가 "collector 에 못
      닿았습니다" 를 띄워 **엉뚱한 층(네트워크·컨테이너)을 의심하게 만든다** — 실제로
      그렇게 한 번 헛짚었다. 카드가 "무슨 파일이 없다"를 그대로 말하게 한다.
    """
    try:
        blocks = _read_blocks()
    except FileNotFoundError:
        out = build_payload({})
        out["note"] = f"원천 파일이 없습니다 — {SRC_PATH}"
        return out
    return build_payload(blocks)
