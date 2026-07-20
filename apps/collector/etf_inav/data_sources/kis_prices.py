from __future__ import annotations

import argparse
import csv
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

from kis_api.auth import KisAuth, KisCredentials
from kis_api.futureoption_master import DomesticFutureOptionMaster
from kis_api.identifier import fetch_ticker_by_isin
from kis_api.master import KisMaster
from kis_api.rest_client import KisRestClient
from kis_api.store import KisStore


DEFAULT_MONITOR_DIR = Path(r"S:\@@운용폴더(중요)\2_Monitoring\Monitoring")
DEFAULT_SHEET_NAME = "GE_PDF"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
NEW_ROOT = Path(__file__).resolve().parents[5]
DEFAULT_STOCK_LIST_OUTPUT_DIR = NEW_ROOT / "data" / "ETF_iNAV모니터" / "output" / "results" / "stock_list"
KST = timezone(timedelta(hours=9))


STOCK_LIST_COLUMNS = ["ISIN", "ticker"]
PRICE_COLUMNS = [
    "ISIN",
    "ticker",
    "exchange",
    "currency",
    "price",
    "base_price",
    "open_price",
    "high_price",
    "low_price",
    "volume",
    "value",
    "trade_time",
    "rec_time",
]


def today_yyyymmdd() -> str:
    return datetime.now(KST).strftime("%Y%m%d")


def default_monitor_path(date_text: str) -> Path:
    return DEFAULT_MONITOR_DIR / f"Global_Monitor_{date_text}.xlsm"


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def ticker_to_symbol(ticker: str) -> str:
    text = clean_text(ticker).upper()
    if text.endswith(" US"):
        return text[:-3].strip()
    return text


def chunked(values, size: int):
    for start in range(0, len(values), size):
        yield values[start : start + size]


def read_stock_rows_from_monitor(
    file_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    isin_col: str = "CI",
    ticker_col: str = "CJ",
    keep_duplicates: bool = False,
) -> list[dict[str, str]]:
    if not file_path.exists():
        raise FileNotFoundError(f"Monitor file not found: {file_path}")

    min_col = column_index_from_string(isin_col)
    max_col = column_index_from_string(ticker_col)
    if min_col > max_col:
        raise ValueError(f"Invalid column range: {isin_col}:{ticker_col}")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        sheet = workbook[sheet_name]
        rows = []
        seen_isins = set()

        for isin_value, ticker_value in sheet.iter_rows(
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            isin = clean_text(isin_value).upper()
            ticker = clean_text(ticker_value).upper()

            if not isin or not ticker:
                continue
            if not ticker.endswith("US"):
                continue
            if not keep_duplicates and isin in seen_isins:
                continue

            rows.append({"ISIN": isin, "ticker": ticker})
            seen_isins.add(isin)

        return rows
    finally:
        workbook.close()


def write_csv(path: Path, rows: list[dict], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_stock_list_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        rows = []
        for row in reader:
            isin = clean_text(row.get("ISIN")).upper()
            ticker = clean_text(row.get("ticker")).upper()
            if isin:
                rows.append({"ISIN": isin, "ticker": ticker})
        return rows


def _now_kst_string() -> str:
    return datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")


def _is_korean_isin(isin: str) -> bool:
    return isin.startswith("KR") and len(isin) == 12


def _is_korean_symbol(value: str) -> bool:
    return len(value) == 6 and value.isalnum()


def _symbol_from_korean_isin(isin: str) -> str:
    if len(isin) >= 9:
        symbol = isin[3:9].upper()
        if _is_korean_symbol(symbol):
            return symbol
    return ""


def _is_stock_future_row(row: dict) -> bool:
    return (
        clean_text(row.get("MKT_ID")).upper() == "DRV"
        and clean_text(row.get("SECUGRP_ID")).upper() == "FU"
    )


def resolve_instruments(
    stock_rows: list[dict[str, str]],
    store: KisStore,
    master: KisMaster,
    openfigi_api_key: str | None,
    verify_ssl: bool,
    timeout: int,
    isin_batch_size: int,
    batch_delay_seconds: float,
    futureoption_master: DomesticFutureOptionMaster | None = None,
) -> tuple[list[dict], list[dict]]:
    """Map ISINs to (exchange, symbol) using OpenFIGI + KIS master.

    Returns (instruments, unresolved).
    """
    overseas_isins: list[str] = []
    input_ticker_by_isin: dict[str, str] = {}
    domestic_instruments: list[dict] = []
    domestic_unresolved: list[dict] = []
    futureoption_master = futureoption_master or DomesticFutureOptionMaster(
        timeout=timeout,
        verify_ssl=verify_ssl,
    )
    for row in stock_rows:
        isin = clean_text(row.get("ISIN")).upper()
        if not isin or isin in input_ticker_by_isin:
            continue
        ticker = clean_text(row.get("ticker")).upper()
        input_ticker_by_isin[isin] = ticker
        if _is_korean_isin(isin) and _is_stock_future_row(row):
            hit = futureoption_master.lookup_stock_future(isin)
            if hit is not None:
                domestic_instruments.append(
                    {
                        "isin": isin,
                        "ticker": hit.short_code,
                        "exchange": "KFO",
                        "name": hit.name or clean_text(row.get("name")),
                        "currency": hit.currency,
                    }
                )
            else:
                domestic_unresolved.append({"ISIN": isin, "ticker": ticker})
        elif _is_korean_isin(isin):
            symbol = ticker if _is_korean_symbol(ticker) else _symbol_from_korean_isin(isin)
            if symbol:
                domestic_instruments.append(
                    {
                        "isin": isin,
                        "ticker": symbol,
                        "exchange": "KRX",
                        "name": clean_text(row.get("name")),
                        "currency": "KRW",
                    }
                )
            else:
                domestic_unresolved.append({"ISIN": isin, "ticker": ticker})
        else:
            overseas_isins.append(isin)

    if not overseas_isins:
        return domestic_instruments, domestic_unresolved

    mappings = fetch_ticker_by_isin(
        overseas_isins,
        store=store,
        openfigi_api_key=openfigi_api_key,
        timeout=timeout,
        verify=verify_ssl,
        batch_size=isin_batch_size,
        batch_delay_seconds=batch_delay_seconds,
    )

    instruments: list[dict] = list(domestic_instruments)
    unresolved: list[dict] = list(domestic_unresolved)

    if not mappings:
        return instruments, unresolved
    for mapping in mappings:
        isin = mapping["isin"]
        ticker = mapping.get("ticker") or ticker_to_symbol(input_ticker_by_isin.get(isin, ""))
        exchange = mapping.get("exchange")
        currency = mapping.get("currency")

        # OpenFIGI must give both a ticker and a KIS-supported exchange.
        if not ticker or not exchange:
            unresolved.append({"ISIN": isin, "ticker": ticker or input_ticker_by_isin.get(isin, "")})
            continue

        # KIS master enhances with the canonical symbol/currency when present;
        # if the master has no entry we still trust OpenFIGI's resolution.
        master_hit = master.lookup(ticker, exchange_hint=exchange)
        if master_hit is not None:
            instruments.append(
                {
                    "isin": isin,
                    "ticker": master_hit.symbol,
                    "exchange": master_hit.code,
                    "name": master_hit.english_name or mapping.get("name") or "",
                    "currency": master_hit.currency or currency,
                }
            )
        else:
            instruments.append(
                {
                    "isin": isin,
                    "ticker": ticker,
                    "exchange": exchange,
                    "name": mapping.get("name") or "",
                    "currency": currency,
                }
            )

    return instruments, unresolved


def fetch_prices_by_isin(
    stock_rows: list[dict[str, str]],
    db_path: Path | None = None,
    verify_ssl: bool = False,
    timeout: int = 10,
    isin_batch_size: int = 10,
    price_batch_delay_seconds: float = 0.05,
    batch_delay_seconds: float = 2.0,
    openfigi_api_key: str | None = None,
    auth: KisAuth | None = None,
    master: KisMaster | None = None,
    paper: bool = False,
) -> dict:
    """Resolve ISIN list to KIS snapshots.

    Returns a dict with keys: instruments, unresolved, prices, rec_time.
    """
    store = KisStore(db_path=db_path)
    store.init_db()

    if auth is None:
        auth = KisAuth(KisCredentials.from_env(paper=paper), verify_ssl=verify_ssl, timeout=timeout)
    if master is None:
        master = KisMaster(verify_ssl=verify_ssl)
    client = KisRestClient(auth=auth, timeout=timeout)

    instruments, unresolved = resolve_instruments(
        stock_rows,
        store=store,
        master=master,
        openfigi_api_key=openfigi_api_key,
        verify_ssl=verify_ssl,
        timeout=timeout,
        isin_batch_size=isin_batch_size,
        batch_delay_seconds=batch_delay_seconds,
    )

    targets = [(row["exchange"], row["ticker"]) for row in instruments]
    snapshots = client.snapshots(targets, batch_delay_seconds=price_batch_delay_seconds)

    rec_time = _now_kst_string()
    snapshot_by_key = {(s["exchange"], s["symbol"]): s for s in snapshots}

    prices: list[dict] = []
    for instrument in instruments:
        key = (instrument["exchange"], instrument["ticker"])
        snap = snapshot_by_key.get(key) or {}
        prices.append(
            {
                "isin": instrument["isin"],
                "ticker": instrument["ticker"],
                "exchange": instrument["exchange"],
                "currency": snap.get("currency") or instrument.get("currency"),
                "last": snap.get("last"),
                "base": snap.get("base"),
                "open": snap.get("open"),
                "high": snap.get("high"),
                "low": snap.get("low"),
                "volume": snap.get("volume"),
                "value": snap.get("value"),
                "trade_time": snap.get("trade_time"),
                "rec_time": rec_time,
                "raw": snap.get("raw"),
                "error": snap.get("error"),
            }
        )

    store.upsert_ticks(
        [
            {
                "exchange": p["exchange"],
                "symbol": p["ticker"],
                "trade_time": p.get("trade_time") or rec_time,
                "observed_at": rec_time,
                "last": p.get("last"),
                "open": p.get("open"),
                "high": p.get("high"),
                "low": p.get("low"),
                "volume": p.get("volume"),
            }
            for p in prices
            if p.get("last") is not None
        ]
    )

    return {
        "instruments": instruments,
        "unresolved": unresolved,
        "prices": prices,
        "rec_time": rec_time,
    }


def build_price_output_rows(prices: list[dict]) -> list[dict]:
    rows = []
    for price in prices:
        rows.append(
            {
                "ISIN": price.get("isin", ""),
                "ticker": price.get("ticker", ""),
                "exchange": price.get("exchange", ""),
                "currency": price.get("currency"),
                "price": price.get("last"),
                "base_price": price.get("base"),
                "open_price": price.get("open"),
                "high_price": price.get("high"),
                "low_price": price.get("low"),
                "volume": price.get("volume"),
                "value": price.get("value"),
                "trade_time": price.get("trade_time"),
                "rec_time": price.get("rec_time"),
            }
        )
    return rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create stock_list_YYYYMMDD.csv from Global_Monitor CI/CJ columns "
            "and fetch KIS prices by ISIN."
        )
    )
    parser.add_argument("--date", default=today_yyyymmdd(), help="YYYYMMDD date. Default: today in KST.")
    parser.add_argument("--file", type=Path, help="Monitor xlsm path. Default: Global_Monitor_DATE.xlsm.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Worksheet name. Default: GE_PDF.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_STOCK_LIST_OUTPUT_DIR, help="Output directory.")
    parser.add_argument("--db-path", type=Path, default=None, help="SQLite DB path for KIS price cache.")
    parser.add_argument("--isin-col", default="CI", help="ISIN Excel column letter. Default: CI.")
    parser.add_argument("--ticker-col", default="CJ", help="Ticker Excel column letter. Default: CJ.")
    parser.add_argument("--keep-duplicates", action="store_true", help="Keep duplicate ISIN rows.")
    parser.add_argument("--no-fetch-prices", action="store_true", help="Only create the stock list CSV.")
    parser.add_argument("--verify-ssl", action="store_true", help="Enable SSL verification for KIS/OpenFIGI requests.")
    parser.add_argument("--timeout", type=int, default=10, help="HTTP timeout seconds.")
    parser.add_argument("--isin-batch-size", type=int, default=10, help="OpenFIGI ISIN mapping batch size.")
    parser.add_argument("--batch-delay-seconds", type=float, default=2.0, help="Delay between OpenFIGI batches.")
    parser.add_argument("--price-batch-delay-seconds", type=float, default=0.05, help="Delay between KIS snapshot calls.")
    parser.add_argument("--openfigi-api-key", default=None, help="Optional OpenFIGI API key.")
    parser.add_argument("--paper", action="store_true", help="Use KIS paper trading endpoint.")
    return parser.parse_args()


def main() -> int:
    start = time.time()
    args = parse_args()
    monitor_path = args.file or default_monitor_path(args.date)

    stock_list_path = args.output_dir / f"stock_list_{args.date}.csv"
    price_output_path = args.output_dir / f"stock_prices_{args.date}.csv"

    if stock_list_path.exists():
        stock_rows = read_stock_list_csv(stock_list_path)
        print(f"Skipped stock list creation: {stock_list_path} already exists ({len(stock_rows)} rows)")
    else:
        stock_rows = read_stock_rows_from_monitor(
            file_path=monitor_path,
            sheet_name=args.sheet,
            isin_col=args.isin_col,
            ticker_col=args.ticker_col,
            keep_duplicates=args.keep_duplicates,
        )
        write_csv(stock_list_path, stock_rows, STOCK_LIST_COLUMNS)
        print(f"Saved stock list: {stock_list_path} ({len(stock_rows)} rows)")

    if args.no_fetch_prices or not stock_rows:
        return 0

    result = fetch_prices_by_isin(
        stock_rows,
        db_path=args.db_path,
        verify_ssl=args.verify_ssl,
        timeout=args.timeout,
        isin_batch_size=args.isin_batch_size,
        price_batch_delay_seconds=args.price_batch_delay_seconds,
        batch_delay_seconds=args.batch_delay_seconds,
        openfigi_api_key=args.openfigi_api_key,
        paper=args.paper,
    )

    write_csv(price_output_path, build_price_output_rows(result["prices"]), PRICE_COLUMNS)

    print(f"Resolved instruments: {len(result['instruments'])}")
    print(f"Unresolved ISINs: {len(result['unresolved'])}")
    print(f"Saved price output: {price_output_path} ({len(result['prices'])} rows)")

    end = time.time()
    print("걸린시간: ", round(end - start, 2), "초")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
