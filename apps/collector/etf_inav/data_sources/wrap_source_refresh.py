"""AICoreTech 전일종가·비중 자동 갱신 (08:50 KST).

운용역 작업파일(``이상 포트폴리오 수익률.xlsx`` 의 ``수익률_breakdown`` 시트)에서
전일종가(T-1)·비중·현금을 읽어, wrap_watchlist 가 읽는
``AICORETECH_PDF.xlsx`` (평문, ticker|name|weight_pct|prev_close|exchange) 로 재기록한다.
→ 다음 wrap cycle(≤15초)에 자동 반영.

streaming 의 ``_wrap_source_refresh_loop`` 가 매일 08:50 호출. iNAV/WS 와 독립.

★ 설계(fail-loud — 하나라도 어긋나면 기록하지 않고 어제 PDF 유지):
  - 비중(I열 '(T-1)일 비중')에 #DIV/0!/비수치가 1개라도 있으면 거부(과소·오산정 방지).
  - 비중 합이 100%(또는 1.0)에서 벗어나면 거부.
  - 기준일(F헤더 날짜)이 없거나 너무 오래(>5일)됐으면 거부.
  - 이미 같은 기준일을 기록했으면 no-op(소스가 갱신될 때만 덮어씀).
  - 출력은 openpyxl 직접 기록(평문) → 사내 DRM 회피. (Excel 저장 금지)

소스 레이아웃(고정 컬럼, 동적 행):
  A=종목명 · B=티커 · F=전일종가(T-1) · I=(T-1)일 비중 · '(현금)' 라벨 행=현금
  데이터는 3행부터, A/B 가 모두 빈 행에서 종료.
"""
from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

from etf_inav.data_sources.wrap_watchlist import WRAP_CONFIG_PATH

# 소스/시트 기본값(config 로 덮어쓸 수 있음).
DEFAULT_SOURCE_PATH = (
    r"S:\GE\Wonjae\02_운용펀드_글로벌\(Wrap) 한국투자 미국 AI코어테크랩"
    r"\4_수익률\이상 포트폴리오 수익률.xlsx"
)
DEFAULT_SHEET = "수익률_breakdown"
OUTPUT_HEADER = ["ticker", "name", "weight_pct", "prev_close", "exchange"]

# 고정 컬럼 인덱스(1-base) — 사용자 확정 레이아웃.
COL_NAME, COL_TICKER, COL_PREV2, COL_PREV, COL_WEIGHT = 1, 2, 5, 6, 9  # A, B, E, F, I
DATA_START_ROW = 3
BASIS_DATE_CELL = (2, COL_PREV)   # F2 = T-1 기준일(datetime)
PREV2_DATE_CELL = (2, COL_PREV2)  # E2 = T-2 기준일(datetime)
# 표 아래 별도 행(A열 라벨). E·F 에 T-2·T-1 USD/KRW(블룸버그 KRW L160) 가 있다.
FX_ROW_LABEL = "USDKRW"
MAX_BASIS_AGE_DAYS = 5           # 기준일이 이보다 오래되면 stale 로 거부
WEIGHT_SUM_TOLERANCE = 1.0       # 상한: 100+1% 까지 허용
WEIGHT_SUM_MIN = 75.0            # 하한(%): 운용역 시트가 리밸런스 중간 상태(합<100%)여도
                                 # 75% 이상이면 적용(2026-07-08 사용자 결정 — MAGS 편입 건)


def _is_num(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _is_cash(name) -> bool:
    return "현금" in str(name or "")


def _load_cfg() -> dict:
    return json.loads(WRAP_CONFIG_PATH.read_text(encoding="utf-8"))


def _portfolio_sources(cfg: dict) -> dict[str, tuple[Path, str, Path]]:
    """config → {portfolio_key: (source_path, sheet, out_pdf_path)}.

    우선순위: ``wrap_sources`` (랩별 소스 시트). 없으면 레거시 ``aicoretech_source``
    하나만 AICoreTech 로 폴백. 출력경로는 ``portfolio_pdfs[key]`` (현재 monitor_v3).
    """
    pdfs = cfg.get("portfolio_pdfs", {})
    sources = dict(cfg.get("wrap_sources", {}))
    if not sources and cfg.get("aicoretech_source"):
        sources = {"AICoreTech": cfg["aicoretech_source"]}
    out: dict[str, tuple[Path, str, Path]] = {}
    for key, src in sources.items():
        if key not in pdfs:
            print(f"[wrap_source] {key}: portfolio_pdfs 경로 없음 — 스킵", file=sys.stderr)
            continue
        out[key] = (
            Path(src.get("path", DEFAULT_SOURCE_PATH)),
            src.get("sheet", DEFAULT_SHEET),
            Path(pdfs[key]),
        )
    return out


def _read_source(src_path: Path, sheet: str):
    """소스 시트 → (rows, basis_date, extra).

    rows  = [{name,ticker,prev2,prev,weight,is_cash}] — 3행부터 A·B 가 모두 빈 행 직전까지.
    extra = {prev2_date, fx_t2, fx_t1} — T-2 기준일과 T-2·T-1 USD/KRW.
            환율 행(``USDKRW``)은 표 아래 빈 행 너머에 있어 그리드 전체를 훑어 찾는다.
    """
    import openpyxl

    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"시트 없음: {sheet}")
    ws = wb[sheet]
    grid = list(ws.iter_rows(values_only=True))
    wb.close()

    def cell(r1: int, c1: int):
        r = grid[r1 - 1] if r1 - 1 < len(grid) else ()
        return r[c1 - 1] if c1 - 1 < len(r) else None

    basis_raw = cell(*BASIS_DATE_CELL)
    basis_date = basis_raw.date() if isinstance(basis_raw, datetime) else None
    prev2_raw = cell(*PREV2_DATE_CELL)
    prev2_date = prev2_raw.date() if isinstance(prev2_raw, datetime) else None

    fx_t2 = fx_t1 = None
    for r1 in range(1, len(grid) + 1):
        if str(cell(r1, COL_NAME) or "").strip() == FX_ROW_LABEL:
            fx_t2, fx_t1 = cell(r1, COL_PREV2), cell(r1, COL_PREV)
            break

    rows = []
    for r1 in range(DATA_START_ROW, len(grid) + 1):
        name = cell(r1, COL_NAME)
        ticker = cell(r1, COL_TICKER)
        if (name in (None, "")) and (ticker in (None, "")):
            break  # 데이터 끝
        if str(name or "").strip() == FX_ROW_LABEL:
            continue  # 환율 행은 보유종목이 아니다(표 안으로 들어와도 방어)
        rows.append({
            "name": name,
            "ticker": "" if ticker is None else str(ticker).strip().upper(),
            "prev2": cell(r1, COL_PREV2),
            "prev": cell(r1, COL_PREV),
            "weight": cell(r1, COL_WEIGHT),
            "is_cash": _is_cash(name),
        })
    extra = {
        "prev2_date": prev2_date,
        "fx_t2": fx_t2 if _is_num(fx_t2) and fx_t2 > 0 else None,
        "fx_t1": fx_t1 if _is_num(fx_t1) and fx_t1 > 0 else None,
    }
    return rows, basis_date, extra


def _write_pdf(out_path: Path, out_rows: list[dict]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(OUTPUT_HEADER)
    for h in out_rows:
        ws.append([h["ticker"], h["name"], h["weight_pct"], h["prev_close"], h["exchange"]])
    tmp = out_path.with_suffix(out_path.suffix + ".tmp")
    wb.save(tmp)
    wb.close()
    os.replace(tmp, out_path)


def refresh_portfolio_pdf(
    label: str, src_path: Path, sheet: str, out_path: Path
) -> tuple[bool, dict]:
    """소스 시트 → 표준 PDF(out_path) 재기록. (ok, info) 반환.

    ok=True 일 때만 파일을 덮어쓴다. 검증 실패/no-op 이면 기존 PDF 를 그대로 둔다.
    CoreTech(수익률_breakdown)·TORUS(TORUS_수익률) 동일 레이아웃이라 공용
    (A=종목명[TORUS는 빈값] · B=티커 · F=전일종가(T-1) · I=(T-1)일 비중 · F2=기준일 · `(현금)`행).
    """
    if not src_path.exists():
        return False, {"reason": "no_source", "path": str(src_path)}

    try:
        rows, basis_date, _extra = _read_source(src_path, sheet)
    except Exception as exc:
        return False, {"reason": f"read_error: {exc}"}

    # ── 기준일(freshness) 가드 ──
    if basis_date is None:
        return False, {"reason": "no_basis_date(F2)"}
    today = datetime.now().date()
    if (today - basis_date).days > MAX_BASIS_AGE_DAYS:
        return False, {"reason": f"stale_basis({basis_date})"}

    sidecar = out_path.with_suffix(out_path.suffix + ".basis")
    prev_basis = sidecar.read_text(encoding="utf-8").strip() if sidecar.exists() else ""
    basis_str = basis_date.strftime("%Y%m%d")
    if prev_basis and prev_basis >= basis_str:
        return False, {"reason": "no_new_basis", "basis": basis_str, "applied": prev_basis}

    # ── 비중 검증(현금 포함). 증권 비중에 #DIV/0!/비수치 1개라도 있으면 거부 ──
    securities = [r for r in rows if not r["is_cash"]]
    cash_rows = [r for r in rows if r["is_cash"]]
    if not securities:
        return False, {"reason": "no_securities"}
    bad = [r["ticker"] or r["name"] for r in securities if not _is_num(r["weight"])]
    if bad:
        return False, {"reason": "non_numeric_weight", "tickers": bad[:5], "n": len(bad)}
    cash_w_raw = sum(r["weight"] for r in cash_rows if _is_num(r["weight"]))
    sec_w_raw = sum(r["weight"] for r in securities)
    total_raw = sec_w_raw + cash_w_raw

    # 스케일 판별: 분수(합 0.75~1.05) → ×100, 퍼센트(합 75~105) → ×1, 그 외 → 거부.
    # 두 범위는 겹치지 않아 스케일 오판 없음.
    if WEIGHT_SUM_MIN / 100.0 <= total_raw <= 1.0 + 0.05:
        factor = 100.0
    elif WEIGHT_SUM_MIN <= total_raw <= 100.0 + 5.0:
        factor = 1.0
    else:
        return False, {"reason": f"weight_sum_off({total_raw:.4f})"}
    final_total = total_raw * factor
    if not (WEIGHT_SUM_MIN <= final_total <= 100.0 + WEIGHT_SUM_TOLERANCE):
        return False, {"reason": f"weight_sum_off({final_total:.3f}%)"}

    # ── 출력행 구성 ──
    out_rows: list[dict] = []
    zero_prev = 0
    for r in securities:
        prev = r["prev"] if (_is_num(r["prev"]) and r["prev"] > 0) else None
        if prev is None:
            zero_prev += 1
        out_rows.append({
            "ticker": r["ticker"],
            "name": "" if r["name"] is None else str(r["name"]),
            "weight_pct": round(r["weight"] * factor, 6),
            "prev_close": prev,
            "exchange": "",  # wrap_watchlist 가 KisMaster 로 자동 해소
        })
    # 현금: ticker=CASH(미해소→0 기여), 비중만 보존
    if cash_rows and _is_num(cash_rows[0]["weight"]):
        out_rows.append({
            "ticker": "CASH",
            "name": "현금",
            "weight_pct": round(cash_w_raw * factor, 6),
            "prev_close": None,
            "exchange": "",
        })

    try:
        _write_pdf(out_path, out_rows)
        sidecar.write_text(basis_str, encoding="utf-8")
    except Exception as exc:
        return False, {"reason": f"write_error: {exc}"}

    info = {
        "basis": basis_str,
        "n_securities": len(securities),
        "cash_pct": round(cash_w_raw * factor, 3),
        "total_pct": round(final_total, 3),
        "zero_prev": zero_prev,
        "out": str(out_path),
    }
    print(f"[wrap_source] {label} 갱신: {info}", file=sys.stderr)
    return True, info


def refresh_all_portfolio_pdfs() -> dict[str, tuple[bool, dict]]:
    """config 의 모든 wrap 포트폴리오 소스 → 각 PDF 재기록. {key: (ok, info)} 반환.

    streaming 의 ``_wrap_source_refresh_loop`` 가 매일 08:50 호출. 한 랩이 실패해도
    다른 랩은 독립적으로 갱신된다(fail-loud per portfolio).
    """
    cfg = _load_cfg()
    results: dict[str, tuple[bool, dict]] = {}
    for key, (src_path, sheet, out_path) in _portfolio_sources(cfg).items():
        try:
            results[key] = refresh_portfolio_pdf(key, src_path, sheet, out_path)
        except Exception as exc:
            results[key] = (False, {"reason": f"error: {exc}"})
    return results


def refresh_aicoretech_pdf() -> tuple[bool, dict]:
    """레거시 호환: AICoreTech 단일 갱신(외부 호출/수동 실행용)."""
    src = _portfolio_sources(_load_cfg()).get("AICoreTech")
    if not src:
        return False, {"reason": "no_aicoretech_source"}
    return refresh_portfolio_pdf("AICoreTech", *src)
