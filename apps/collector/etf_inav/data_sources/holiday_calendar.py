"""Market holiday calendar backed by a shared Excel sheet.

One workbook per year lives at::

    S:\\GE\\raw\\data\\_공유모듈\\Holidays\\Holidays_{YYYY}.xlsx  (2026-07-10 이전; 구 06_퀀트전략\\@데이터관리\\Holidays)

Schema (one row per *calendar* day — weekends included):

    일자 | 한국 | 일본 | 대만 | 홍콩 | 중국 | 미국 | 영국 | 비고

Each country cell is ``0`` (개장/open) or ``1`` (휴장/closed). ``비고`` holds the
free-text reason when any country is closed that day.

The reader maps each country to the KIS exchange codes the iNAV engine uses, so
callers can ask "which exchanges are closed today" and force those components to
their previous-day close (base price) instead of a stale live tick.

Fail-open policy: when the year file or the day's row is missing, every market is
treated as open and a one-time warning is logged. This never blocks iNAV.

Weekend assumption: weekends are NOT special-cased here — the sheet is the single
source of truth, so weekend rows must carry ``1`` for markets that do not trade
on weekends (all markets listed above are Mon–Fri).
"""

from __future__ import annotations

import sys
from datetime import date as dt_date, datetime, timedelta, timezone
from pathlib import Path

KST = timezone(timedelta(hours=9))

# 공유 휴장일 캘린더 SSOT — 06\@데이터관리\Holidays 이전 완료(2026-07-10) → _공유모듈\Holidays.
# ★HOL-CUTOVER: 현재 06 스냅샷(byte 동일). 연1회 수기 갱신 원본은 담당자 결정.
DEFAULT_HOLIDAY_DIR = Path(r"S:\GE\raw\data\_공유모듈\Holidays")
FILENAME_PATTERN = "Holidays_{year}.xlsx"

DATE_HEADER = "일자"
NOTE_HEADER = "비고"

# Excel country column -> KIS exchange codes seen in the iNAV components.
# 순서 = 엑셀 컬럼 순서 = 대시보드 배지/hover 표시 순서.
#
# 영국(LSE)은 forward-compatibility 용(현재 UK 보유 없음). 스위스/호주/캐나다는
# 휴장 '표시'(엑셀·배지·hover)용으로만 추가한 국가라 거래소 매핑이 비어 있다:
#   - 스위스/호주: KIS 미지원(마스터·EXCD 코드 없음).
#   - 캐나다: 캐나다 본사 종목도 ISIN(CA) -> 미국 거래소(NAS/NYS/AMS)로 거래되므로
#     '미국' 휴장을 따른다. 여기에 미국 코드를 넣으면 캐나다 단독 휴장일(예: 캐나다
#     데이)에 미국 종목까지 잘못 base price 로 폴백되므로 비워 둔다.
COUNTRY_TO_EXCHANGES: dict[str, set[str]] = {
    "한국": {"KRX", "KFO"},
    "미국": {"NAS", "NYS", "AMS", "BAQ", "BAY", "BAA"},
    "홍콩": {"HKS"},
    "중국": {"SHS", "SZS"},
    "일본": {"TSE"},
    "대만": {"TWSE", "TPEX"},
    "영국": {"LSE"},
    "스위스": set(),
    "호주": set(),
    "캐나다": set(),
}
COUNTRY_COLUMNS = list(COUNTRY_TO_EXCHANGES.keys())


def _coerce_date(value) -> dt_date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, dt_date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _is_closed_value(value) -> bool:
    """True only when the cell clearly means 휴장 (1)."""
    if value is None:
        return False
    try:
        return int(float(str(value).strip())) == 1
    except (TypeError, ValueError):
        return False


class HolidayCalendar:
    def __init__(
        self,
        base_dir: Path = DEFAULT_HOLIDAY_DIR,
        filename_pattern: str = FILENAME_PATTERN,
    ):
        self.base_dir = Path(base_dir)
        self.filename_pattern = filename_pattern
        # year -> (file_mtime, {date: {country: closed_bool}})
        self._cache: dict[int, tuple[float, dict[dt_date, dict[str, bool]]]] = {}
        self._warned: set[str] = set()

    def _warn_once(self, key: str, message: str) -> None:
        if key not in self._warned:
            self._warned.add(key)
            print(f"[holiday] {message}", file=sys.stderr)

    def _path_for_year(self, year: int) -> Path:
        return self.base_dir / self.filename_pattern.format(year=year)

    def _load_year(self, year: int) -> dict[dt_date, dict[str, bool]]:
        path = self._path_for_year(year)
        if not path.exists():
            self._warn_once(
                f"missing-file:{year}",
                f"file not found: {path} → 전체 개장으로 간주(fail-open)",
            )
            return {}
        try:
            mtime = path.stat().st_mtime
        except OSError:
            mtime = 0.0
        cached = self._cache.get(year)
        if cached is not None and cached[0] == mtime:
            return cached[1]
        try:
            table = self._parse(path)
        except Exception as exc:  # noqa: BLE001 - never let a bad sheet block iNAV
            self._warn_once(
                f"parse-error:{year}",
                f"failed to read {path}: {exc!r} → 전체 개장으로 간주(fail-open)",
            )
            table = {}
        self._cache[year] = (mtime, table)
        return table

    def _parse(self, path: Path) -> dict[dt_date, dict[str, bool]]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                self._warn_once(f"empty:{path}", f"empty sheet: {path}")
                return {}

            header_index = {
                str(name).strip(): idx
                for idx, name in enumerate(header)
                if name is not None
            }
            date_idx = header_index.get(DATE_HEADER, 0)
            country_idx = {
                country: header_index[country]
                for country in COUNTRY_COLUMNS
                if country in header_index
            }
            missing = [c for c in COUNTRY_COLUMNS if c not in country_idx]
            if missing:
                self._warn_once(
                    f"missing-cols:{path}",
                    f"columns absent in {path.name}: {missing} (해당국은 개장 처리)",
                )

            table: dict[dt_date, dict[str, bool]] = {}
            for row in rows:
                if not row or date_idx >= len(row):
                    continue
                day = _coerce_date(row[date_idx])
                if day is None:
                    continue
                table[day] = {
                    country: _is_closed_value(row[idx]) if idx < len(row) else False
                    for country, idx in country_idx.items()
                }
            return table
        finally:
            workbook.close()

    def _as_date(self, on) -> dt_date:
        if on is None:
            return datetime.now(KST).date()
        coerced = _coerce_date(on)
        return coerced if coerced is not None else datetime.now(KST).date()

    def closed_markets(self, on=None) -> set[str]:
        """Country names that are 휴장 on the given date (default: today KST)."""
        day = self._as_date(on)
        table = self._load_year(day.year)
        record = table.get(day)
        if record is None:
            if table:  # file loaded but this date is absent → fail-open + warn
                self._warn_once(
                    f"missing-date:{day}",
                    f"{day} 행이 없음 → 전체 개장으로 간주(fail-open)",
                )
            return set()
        return {country for country, closed in record.items() if closed}

    def closed_exchanges(self, on=None) -> set[str]:
        """KIS exchange codes whose market is 휴장 on the given date."""
        exchanges: set[str] = set()
        for country in self.closed_markets(on):
            exchanges |= COUNTRY_TO_EXCHANGES.get(country, set())
        return exchanges

    def is_market_closed(self, country: str, on=None) -> bool:
        return country in self.closed_markets(on)

    def us_status(self, on=None) -> bool | None:
        """미국 휴장 여부. True=휴장, False=개장, None=데이터 없음(호출측 폴백 가능)."""
        day = self._as_date(on)
        table = self._load_year(day.year)
        record = table.get(day)
        if record is None or "미국" not in record:
            return None
        return record["미국"]

    def overview(self, on=None) -> dict:
        """7개국 종합 휴장 현황 + 해당 날짜 데이터 존재 여부.

        대시보드 상태 배지(정상영업일/일부휴장/휴장일/데이터오류)와 hover 테이블의
        단일 소스. ``has_data`` 가 False 면 그 날짜 행을 엑셀에서 못 읽은 것이므로
        (파일/행/파싱 실패) 호출측이 '데이터오류'로 표시할 수 있다.

        return::

            {
                "date": dt_date,
                "weekend": bool,            # 토/일
                "has_data": bool,           # 엑셀에서 그 날짜 행을 실제로 읽었는지
                "countries": {국가: 휴장bool}  # COUNTRY_COLUMNS 순서, True=휴장
            }
        """
        day = self._as_date(on)
        table = self._load_year(day.year)
        record = table.get(day)
        rec = record or {}
        return {
            "date": day,
            "weekend": day.weekday() >= 5,
            "has_data": record is not None,
            "countries": {c: bool(rec.get(c, False)) for c in COUNTRY_COLUMNS},
        }


# Module-level default instance + thin convenience wrappers.
_default = HolidayCalendar()


def closed_markets(on=None) -> set[str]:
    return _default.closed_markets(on)


def closed_exchanges(on=None) -> set[str]:
    return _default.closed_exchanges(on)


def is_market_closed(country: str, on=None) -> bool:
    return _default.is_market_closed(country, on)


def us_status(on=None) -> bool | None:
    return _default.us_status(on)


def overview(on=None) -> dict:
    return _default.overview(on)
