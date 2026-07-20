# -*- coding: utf-8 -*-
"""WRAP 종목 분류(대/중/소분류) 추출기.

운용역 소스 엑셀(``이상 포트폴리오 수익률.xlsx``)의 ``종목_분류`` 시트
(종목 | 티커 | 대분류 | 중분류 | 소분류)를 읽어, ticker→분류 사전을
SSOT data 폴더의 ``classification.json`` 으로 평문 기록한다.

설계:
  - 매일 08:00 wrap_source_refresh 루프에서 PDF 재기록과 함께 1회 추출(독립 try/except).
  - wrap_watchlist.run_cycle 가 ``load_classification`` 으로 캐시 로드(파일 mtime 변경 시 재적재)
    하여 wrap.js 보유종목에 cat1/cat2/cat3 를 실어 보낸다.
  - 조인 키 = ticker.upper() (PDF 티커와 분류표 티커가 동일 포맷: BE/SNDK/INTC…).
  - 소스가 Excel 에서 열려 잠겨 있어도 공유읽기 사본으로 우회해 읽는다.
"""
from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile
from pathlib import Path

# 시트 헤더(한글) → 표준 키. 헤더 매칭 실패 시 위치(0~4)로 폴백.
_HEADER_MAP = {
    "종목": "name",
    "티커": "ticker",
    "대분류": "cat1",
    "중분류": "cat2",
    "소분류": "cat3",
}
_POS_KEYS = ["name", "ticker", "cat1", "cat2", "cat3"]


def _open_workbook_shared(path: Path):
    """openpyxl 로 read_only 로드. 잠금/공유위반 시 임시 사본으로 우회."""
    import openpyxl

    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True), None
    except (PermissionError, OSError):
        # Excel 점유 등으로 직접 열기 실패 → 공유읽기 사본 후 로드
        fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix="wrap_cls_")
        os.close(fd)
        with open(path, "rb") as src, open(tmp, "wb") as dst:
            shutil.copyfileobj(src, dst)
        return openpyxl.load_workbook(tmp, read_only=True, data_only=True), tmp


def _read_classification_sheet(path: Path, sheet: str) -> dict[str, dict]:
    wb, tmp = _open_workbook_shared(path)
    try:
        if sheet not in wb.sheetnames:
            raise KeyError(f"시트 '{sheet}' 없음 (가능: {wb.sheetnames})")
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
        if tmp:
            try:
                os.remove(tmp)
            except OSError:
                pass

    if not rows:
        return {}
    header = [(str(h).strip() if h is not None else "") for h in rows[0]]
    # 헤더 → 컬럼 인덱스 매핑(한글 헤더 우선, 없으면 위치 폴백)
    col: dict[str, int] = {}
    for i, h in enumerate(header):
        key = _HEADER_MAP.get(h)
        if key and key not in col:
            col[key] = i
    if "ticker" not in col:  # 헤더 매칭 실패 → 위치 기반
        col = {k: i for i, k in enumerate(_POS_KEYS)}

    def cell(r, key):
        i = col.get(key)
        if i is None or i >= len(r):
            return ""
        v = r[i]
        return "" if v is None else str(v).strip()

    out: dict[str, dict] = {}
    for r in rows[1:]:
        if not r:
            continue
        tk = cell(r, "ticker").upper()
        if not tk or not tk.isascii():
            continue
        out[tk] = {
            "name": cell(r, "name"),
            "cat1": cell(r, "cat1"),
            "cat2": cell(r, "cat2"),
            "cat3": cell(r, "cat3"),
        }
    return out


def build_classification_json(
    source_path: Path, sheet: str, out_json: Path
) -> tuple[bool, dict]:
    """``종목_분류`` 시트 → ticker→분류 사전 JSON 을 원자적으로 기록. (ok, info) 반환.

    1종목도 못 읽으면(빈 결과/예외) 기존 JSON 을 보존하고 ok=False.
    """
    source_path, out_json = Path(source_path), Path(out_json)
    if not source_path.exists():
        return False, {"reason": "no_source", "path": str(source_path)}
    try:
        mapping = _read_classification_sheet(source_path, sheet)
    except Exception as exc:
        return False, {"reason": f"read_error: {exc}"}
    if not mapping:
        return False, {"reason": "empty"}

    payload = {"source": source_path.name, "sheet": sheet, "tickers": mapping}
    out_json.parent.mkdir(parents=True, exist_ok=True)
    tmp = out_json.with_suffix(out_json.suffix + ".tmp")
    tmp.write_text(
        json.dumps(payload, ensure_ascii=False, indent=0), encoding="utf-8"
    )
    os.replace(tmp, out_json)
    return True, {"count": len(mapping), "out": str(out_json)}


# ── wrap_watchlist 용 캐시 로더 ──────────────────────────────────────────
_CACHE: dict[str, object] = {"path": None, "mtime": None, "data": {}}


def load_classification(out_json: Path) -> dict[str, dict]:
    """classification.json 로드. 파일 mtime 이 바뀔 때만 재적재(매 사이클 호출 가벼움)."""
    out_json = Path(out_json)
    if not out_json.exists():
        return {}
    try:
        mtime = out_json.stat().st_mtime
    except OSError:
        return _CACHE.get("data", {})  # type: ignore[return-value]
    if _CACHE["path"] == str(out_json) and _CACHE["mtime"] == mtime:
        return _CACHE["data"]  # type: ignore[return-value]
    try:
        payload = json.loads(out_json.read_text(encoding="utf-8"))
        data = payload.get("tickers", {}) if isinstance(payload, dict) else {}
    except Exception as exc:
        print(f"[wrap_classification] load 실패: {exc}", file=sys.stderr)
        return _CACHE.get("data", {})  # type: ignore[return-value]
    _CACHE.update(path=str(out_json), mtime=mtime, data=data)
    return data
